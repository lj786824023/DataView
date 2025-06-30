import os
import re
import sys
from datetime import datetime
from operator import itemgetter
from time import sleep
from loguru import logger
import chardet
from PySide6 import QtWidgets
from PySide6.QtCore import Qt, QTimer
from PySide6.QtGui import QIcon, QColor
from PySide6.QtWidgets import QMainWindow, QMessageBox, QFileDialog, QTableWidgetItem, QTableWidget, QWidget
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from qfluentwidgets import InfoBar, InfoBarPosition, IndeterminateProgressRing, FluentIcon, ColorDialog, qconfig, \
    RoundMenu, Action, MessageBox
import ColumnTrans
import FileBlood
import Similarity
import UpdateLog
from ConfigFile import ConfigFile
from Func import pars_text
from Highlighter import Highlighter
from MyThread import PoolThread, DBQueryThread, SilenceThread
from ui.MainUI import Ui_MainWindow


class MyMainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.myUI = Ui_MainWindow()
        self.myUI.setupUi(self)
        # self.setWindowFlag(QtCore.Qt.FramelessWindowHint) # 设置窗口无边框
        # self.setAttribute(QtCore.Qt.WA_TranslucentBackground) # 设置窗口的背景为半透明
        self.myUI.tbw_table.geometry()
        # 初始化部分参数
        self.cf_db = ConfigFile(os.path.dirname(sys.argv[0]) + "/_internal/aaa_etc/db_config.ini", "utf=8")  # 配置文件读取器
        self.thread_get_tables = None  # 表线程
        self.thread_get_tab_datas = None  # 表数据线程
        self.thread_get_procedures = None  # 过程线程
        self.thread_get_gbase_sql = None  # 过程线程
        self.thread_sql_body_1 = None  # sql线程
        self.thread_sql_body_2 = None  # sql线程
        self.thread_sql_body_3 = None  # sql线程
        self.db_info = None  # 当前连接信息
        self.tables = [("pdm", "table_list", "")]  # 所有表
        self.procedures = [("pdm", "procedure", "")]  # 所有存储过程
        self.sql_book_dir = "_internal/aaa_book/"  # sql书签目录
        self.highlighter_sql_body_1 = Highlighter(self.myUI.edt_sql_body_1.document())  # sql编辑框高亮
        self.highlighter_sql_body_2 = Highlighter(self.myUI.edt_sql_body_2.document())  # sql编辑框高亮
        self.highlighter_sql_body_3 = Highlighter(self.myUI.edt_sql_body_3.document())  # sql编辑框高亮
        self.highlighter_procdure_body = Highlighter(self.myUI.edt_procedure_body.document())  # 过程体高亮
        self.tbw_current = None  # 当前选中的tbw
        self.ui_col_trans = ColumnTrans.MyMainWindow()  # 字段翻译
        self.ui_file_blood = FileBlood.MyMainWindow()  # 文件血缘
        self.ui_similarity = Similarity.MyMainWindow()  # 相似度窗口
        self.ui_update_log = UpdateLog.MyMainWindow()  # 更新日志

        self.thread_pool = None  # 当前连接线程池
        self.thread_test_pool = None  # 测试连接线程
        self.thread_table = None  # 表页线程
        self.thread_procedure = None  # 过程页线程
        self.thread_sql_execute_1 = None  # sql执行线程
        self.thread_sql_execute_2 = None  # sql执行线程
        self.thread_sql_execute_3 = None  # sql执行线程

        # 绑定函数
        self.myUI.cbb_choose_db.currentTextChanged.connect(lambda: self.myUI.btn_connect.setText("连接"))
        self.myUI.lswDB.itemClicked.connect(self.show_db_info)  # 数据配置项目点击
        self.myUI.btn_save_db.clicked.connect(self.save_db_info)  # 数据库配置保存按钮点击
        self.myUI.btn_dbinfo_add.clicked.connect(self.add_db_info)  # 数据库配置新增按钮点击
        self.myUI.btn_dbinfo_del.clicked.connect(self.delete_db_info)  # 数据库配置删除按钮点击
        self.myUI.btn_test_db.clicked.connect(self.test_db_connect)  # 数据库配置测试按钮点击
        self.myUI.btn_connect.clicked.connect(self.connect_db)  # 数据库连接按钮点击
        self.myUI.edt_table.textChanged.connect(self.find_tables)  # 表筛选框文本改变
        self.myUI.cbb_database.currentTextChanged.connect(self.find_tables)  # 表筛选下拉框当前文本改变
        self.myUI.edt_find_procedure.textChanged.connect(self.find_procedures)  # 过程筛选框文本变更
        self.myUI.cbb_find_procedure.currentTextChanged.connect(self.find_procedures)  # 过程筛选下拉框当前文本变更
        self.myUI.tbw_table.itemClicked.connect(self.get_columns)  # 表清单项目点击
        self.myUI.tbw_procedure.itemClicked.connect(self.get_procedure_body)  # 过程列表项目点击
        self.myUI.edt_procedure_body.textChanged.connect(self.show_procedure_blood)  # 过程体文本改变
        self.myUI.btn_yf.clicked.connect(self.get_gbase_sql)  # 翻译GBase语句按钮点击
        self.myUI.btn_reset.clicked.connect(self.clear_filter)  # 表重置筛选条件按钮点击
        self.myUI.btn_get_data.clicked.connect(self.get_tab_datas)  # 表查询数据按钮点击
        self.myUI.btn_sql_save_1.clicked.connect(self.save_sql_book)  # sql书签保存按钮点击
        self.myUI.btn_sql_save_2.clicked.connect(self.save_sql_book)  # sql书签保存按钮点击
        self.myUI.btn_sql_save_3.clicked.connect(self.save_sql_book)  # sql书签保存按钮点击
        self.myUI.btn_sql_del_1.clicked.connect(self.del_sql_book)  # sql书签删除按钮点击
        self.myUI.btn_sql_del_2.clicked.connect(self.del_sql_book)  # sql书签删除按钮点击
        self.myUI.btn_sql_del_3.clicked.connect(self.del_sql_book)  # sql书签删除按钮点击
        self.myUI.lsw_book.itemClicked.connect(self.show_sql_body)  # sql书签项目点击
        self.myUI.btn_sql_execute_1.clicked.connect(self.execute_sql_body_1)  # sql执行按钮点击
        self.myUI.btn_sql_execute_2.clicked.connect(self.execute_sql_body_2)  # sql执行按钮点击
        self.myUI.btn_sql_execute_3.clicked.connect(self.execute_sql_body_3)  # sql执行按钮点击
        # self.myUI.btn_mapping.clicked.connect(self.export_mapping)  # mapping导出
        self.myUI.tbw_column.btn_export.clicked.connect(
            lambda: self.export_excel(self.myUI.tbw_column.tbw_table))  # excel导出
        self.myUI.tbw_sql_result_1.btn_export.clicked.connect(
            lambda: self.export_excel(self.myUI.tbw_sql_result_1.tbw_table))
        self.myUI.tbw_sql_result_2.btn_export.clicked.connect(
            lambda: self.export_excel(self.myUI.tbw_sql_result_2.tbw_table))
        self.myUI.tbw_sql_result_3.btn_export.clicked.connect(
            lambda: self.export_excel(self.myUI.tbw_sql_result_3.tbw_table))
        self.myUI.tbw_table.itemSelectionChanged.connect(self.set_tbw_current)  # 表清单项目选择改变
        self.myUI.tbw_column.tbw_table.itemSelectionChanged.connect(self.set_tbw_current)  # 字段表格项目选择改变
        self.myUI.cbb_dpi.currentTextChanged.connect(  # dpi
            lambda x: self.cf_dpi.update("environment", {"QT_FONT_DPI": f"{x.replace('%', '')}"}))
        self.myUI.cbb_dpi.currentTextChanged.connect(lambda: self.createSuccessInfoBar("重启后生效"))

        self.myUI.btn_col_trans.clicked.connect(self.ui_col_trans.show)  # 字段翻译
        self.myUI.btn_file_blood.clicked.connect(self.ui_file_blood.show)  # 字段翻译
        self.myUI.btn_similarity.clicked.connect(self.ui_similarity.show)  # 相似度
        self.myUI.btn_update_log.clicked.connect(self.ui_update_log.show)  # 更新日志

        # 初始化部分控件
        self.init_ui()

    def add_db_info(self):
        """新增数据库连接"""
        db_name = '新建连接'
        db_list = []
        # 获取当前所有连接名
        for i in range(self.myUI.lswDB.count()):
            item = self.myUI.lswDB.item(i).text()
            db_list.append(item)
        # 遍历，对新建连接重命名
        for i in range(1, 100):
            if db_name in db_list:
                db_name = f"新建连接({i})"
            else:
                break

        # 初始化一个空连接
        data_dict = {"name": db_name,
                     "desc": "",
                     "type": "",
                     "host": "",
                     "port": "",
                     "database": "",
                     "username": "",
                     "password": "",
                     "charset": ""}
        # 保存
        self.cf_db.update(db_name, data_dict)
        # 添加项目
        self.myUI.lswDB.addItem(db_name)
        # 设置当前选择项
        self.myUI.lswDB.setCurrentRow(self.myUI.lswDB.count() - 1)
        # 显示到编辑框
        self.myUI.edt_name.setText(db_name)
        self.myUI.edt_desc.setText("")
        self.myUI.cbb_type.setCurrentIndex(0)
        self.myUI.edt_host.setText("")
        self.myUI.edt_port.setText("")
        self.myUI.edt_database.setText("")
        self.myUI.edt_username.setText("")
        self.myUI.edt_password.setText("")
        self.myUI.edt_charset.setText("")
        # 添加到选择框
        self.myUI.cbb_choose_db.addItem(db_name)

        # 初始化状态信息
        # self.myUI.lab_db_test.setText("")
        # self.myUI.lab_db_save.setText("")

    def clear_filter(self):
        """重置筛选条件"""
        # 重置筛选条件下拉框
        self.myUI.cbb_find_col_1.setCurrentIndex(0)
        self.myUI.cbb_find_col_2.setCurrentIndex(0)
        self.myUI.cbb_find_col_3.setCurrentIndex(0)
        # 清空筛选内容
        self.myUI.edt_find_str_1.clear()
        self.myUI.edt_find_str_2.clear()
        self.myUI.edt_find_str_3.clear()

    def check_enable(self):
        """检查工具是否到期"""
        current_date = datetime.now().date()  # 获取当前日期
        enable_date = datetime.strptime("2025-12-31", "%Y-%m-%d").date()  # 工具有效期
        if current_date > enable_date:
            # QMessageBox.information(None, "消息", "已到使用有效期！", QMessageBox.Close)
            w = MessageBox("消息", "已到使用有效期！", self)
            if w.exec():
                self.close()
                sys.exit()
            else:
                self.close()
                sys.exit()

    def connect_db(self):
        """连接数据库查询"""
        # 如果索引为0，则不做操作
        if self.myUI.cbb_choose_db.currentIndex() == -1:
            self.createInfoInfoBar("未选择数据库连接")
            return

        # 重载部分界面
        self.myUI.cbb_database.clear()
        self.myUI.cbb_database.addItems(("所有库", "当前库"))
        self.myUI.edt_table.setText("")
        self.load_tablewidget(self.myUI.tbw_table, head=["库", "表名", "注释"])

        # 表页
        self.myUI.cbb_find_col_1.clear()
        self.myUI.cbb_find_col_2.clear()
        self.myUI.cbb_find_col_3.clear()
        self.myUI.cbb_find_col_1.addItem("--条件1--")
        self.myUI.cbb_find_col_2.addItem("--条件2--")
        self.myUI.cbb_find_col_3.addItem("--条件3--")
        self.myUI.edt_find_str_1.setText("")
        self.myUI.edt_find_str_2.setText("")
        self.myUI.edt_find_str_3.setText("")
        self.myUI.edt_result_row.setText("")
        self.myUI.lab_table_comment.setText("")
        self.myUI.tbw_column.tbw_table.head = ["表名", "表注释", "序号", "字段名", "类型", "字段注释"]
        self.myUI.tbw_column.tbw_table.data = []
        self.myUI.tbw_column.setCur_page(1)
        # 过程页面
        self.myUI.cbb_find_procedure.clear()
        self.myUI.cbb_find_procedure.addItems(("所有库", "当前库"))
        self.myUI.edt_find_procedure.setText("")
        self.load_tablewidget(self.myUI.tbw_procedure, head=["库", "类型", "过程名"])
        self.myUI.lab_procedure_name.setText("")
        self.myUI.edt_procedure_body.setText("")
        self.myUI.edt_procedure_blood.clear()

        # 设置按钮
        self.myUI.btn_connect.setEnabled(False)
        db_name = self.myUI.cbb_choose_db.currentText()  # 获取连接名
        self.db_info = self.cf_db.read()[db_name]  # 读取数据库配置文件获取db信息

        # 创建线程池
        cfg = {'type': self.db_info["type"].lower(),
               'host': self.db_info["host"],
               'port': int(self.db_info["port"]),
               'user': self.db_info["username"],
               'password': self.db_info["password"],
               'schema': self.db_info["database"]}
        self.thread_pool = PoolThread(**cfg)

        # 信号绑定函数
        self.thread_pool.sign_end.connect(lambda: self.createSuccessInfoBar("连接成功"))
        self.thread_pool.sign_end.connect(lambda: self.myUI.btn_connect.setEnabled(True))
        self.thread_pool.sign_end.connect(lambda: self.myUI.btn_connect.setText("已连接"))

        self.thread_pool.sign_end.connect(self.init_connection)  # 初始化连接
        self.thread_pool.sign_end.connect(self.get_schemas)  # 加载schema
        self.thread_pool.sign_end.connect(self.get_tables)  # 加载表清单
        self.thread_pool.sign_end.connect(self.get_procedures)  # 加载过程清单

        self.thread_pool.sign_err.connect(self.createErrorInfoBar)
        self.thread_pool.sign_err.connect(logger.warning)
        self.thread_pool.sign_err.connect(lambda: self.myUI.btn_connect.setEnabled(True))

        # 10秒连接超时
        QTimer.singleShot(10000, lambda: self.timer_out(self.thread_pool, "连接超时"))
        self.thread_pool.start()

    def createErrorInfoBar(self, text):
        w = InfoBar.error(
            title='ERROR',
            content=text,
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.BOTTOM,
            duration=3000,
            parent=self
        )
        # w.addWidget(PushButton('Action'))
        w.show()

    def createInfoInfoBar(self, text):
        w = InfoBar.info(
            title='INFO',
            content=text,
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.BOTTOM,
            duration=3000,
            parent=self
        )
        # w.addWidget(PushButton('Action'))
        w.show()

    def createSuccessInfoBar(self, text):
        InfoBar.success(
            title='SUCCESS',
            content=text,
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.BOTTOM,
            # position='Custom',   # NOTE: use custom info bar manager
            duration=3000,
            parent=self
        )

    def del_sql_book(self):
        """删除书签"""
        if not self.myUI.lsw_book.selectedItems():
            self.createInfoInfoBar("没有选中任何书签")
            return
        # 删除文件
        file_name = self.sql_book_dir + self.myUI.lsw_book.currentItem().text() + ".txt"
        os.remove(file_name)
        # 移除标签
        self.myUI.lsw_book.takeItem(self.myUI.lsw_book.row(self.myUI.lsw_book.currentItem()))

    def delete_db_info(self):
        """删除数据库连接"""
        # 若果没有选中，则不做操作
        if not self.myUI.lswDB.currentItem():
            self.createInfoInfoBar("未选中连接")
            return
        db_name = self.myUI.lswDB.currentItem().text()

        # 从配置文件删除
        self.cf_db.delete(db_name)
        # 从显示框里移除
        self.myUI.lswDB.takeItem(self.myUI.lswDB.row(self.myUI.lswDB.currentItem()))
        # 从选择连接框里移除
        for i in range(self.myUI.cbb_choose_db.count()):
            if self.myUI.cbb_choose_db.itemText(i) == db_name:
                self.myUI.cbb_choose_db.removeItem(i)
                break
        self.createSuccessInfoBar("已删除")

    def execute_sql_body_1(self):
        """处理sql执行按钮"""
        # 如果没有选中数据库，则不操作
        if not self.db_info:
            self.createInfoInfoBar("未连接数据库")
            return
        # 执行前的初始化ui
        self.myUI.edt_sql_log_1.clear()  # 清空当前sql日志
        self.myUI.tbw_sql_result_1.tbw_table.setRowCount(0)
        self.myUI.tbw_sql_result_1.tbw_table.setColumnCount(0)

        self.ipr_sql_1 = IndeterminateProgressRing(self.myUI.tbw_sql_result_1)
        self.ipr_sql_1.move((self.myUI.tbw_sql_result_1.width() - self.ipr_sql_1.width()) / 2,
                            (self.myUI.tbw_sql_result_1.height() - self.ipr_sql_1.height()) / 2)
        self.ipr_sql_1.show()

        sql = self.myUI.edt_sql_body_1.textCursor().selectedText().replace("\u2029",
                                                                           "\n") or self.myUI.edt_sql_body_1.toPlainText()  # 获取选中文本
        row = int(self.myUI.edt_sql_row_1.text()) if self.myUI.edt_sql_row_1.text() else 1000

        self.thread_sql_execute_1 = DBQueryThread(self.connection_sql_1,
                                                  sql=sql,
                                                  row_max=row,
                                                  obj=self.myUI.btn_sql_execute_1.objectName())
        self.thread_sql_execute_1.sign_end.connect(self.get_db_result)
        self.thread_sql_execute_1.sign_end.connect(lambda: self.ipr_sql_1.stop())  # 成功停止进度条
        self.thread_sql_execute_1.sign_err.connect(self.createErrorInfoBar)
        self.thread_sql_execute_1.sign_err.connect(self.myUI.edt_sql_log_1.setText)  # 打印日志
        self.thread_sql_execute_1.sign_err.connect(lambda: self.myUI.btn_sql_execute_1.setEnabled(True))
        self.thread_sql_execute_1.sign_err.connect(lambda: self.ipr_sql_1.stop())  # 失败停止进度条

        # 设置状态
        self.myUI.btn_sql_execute_1.setEnabled(False)
        self.thread_sql_execute_1.start()

    def execute_sql_body_2(self):
        """处理sql执行按钮"""
        # 如果没有选中数据库，则不操作
        if not self.db_info:
            self.createInfoInfoBar("未连接数据库")
            return
        # 执行前的初始化ui
        self.myUI.edt_sql_log_2.clear()  # 清空当前sql日志
        self.myUI.tbw_sql_result_2.tbw_table.setRowCount(0)
        self.myUI.tbw_sql_result_2.tbw_table.setColumnCount(0)

        self.ipr_sql_2 = IndeterminateProgressRing(self.myUI.tbw_sql_result_2)
        self.ipr_sql_2.move((self.myUI.tbw_sql_result_2.width() - self.ipr_sql_2.width()) / 2,
                            (self.myUI.tbw_sql_result_2.height() - self.ipr_sql_2.height()) / 2)
        self.ipr_sql_2.show()

        sql = self.myUI.edt_sql_body_2.textCursor().selectedText().replace("\u2029",
                                                                           "\n") or self.myUI.edt_sql_body_2.toPlainText()  # 获取选中文本
        row = int(self.myUI.edt_sql_row_2.text()) if self.myUI.edt_sql_row_2.text() else 1000

        self.thread_sql_execute_2 = DBQueryThread(self.connection_sql_2,
                                                  sql=sql,
                                                  row_max=row,
                                                  obj=self.myUI.btn_sql_execute_2.objectName())
        self.thread_sql_execute_2.sign_end.connect(self.get_db_result)
        self.thread_sql_execute_2.sign_end.connect(lambda: self.ipr_sql_2.stop())
        self.thread_sql_execute_2.sign_err.connect(self.createErrorInfoBar)
        self.thread_sql_execute_2.sign_err.connect(self.myUI.edt_sql_log_2.setText)
        self.thread_sql_execute_2.sign_err.connect(lambda x: self.myUI.btn_sql_execute_2.setEnabled(True))
        self.thread_sql_execute_2.sign_err.connect(lambda: self.ipr_sql_2.stop())

        # 设置状态
        self.myUI.btn_sql_execute_2.setEnabled(False)
        self.thread_sql_execute_2.start()

    def execute_sql_body_3(self):
        """处理sql执行按钮"""
        # 如果没有选中数据库，则不操作
        if not self.db_info:
            self.createInfoInfoBar("未连接数据库")
            return
        # 执行前的初始化ui
        self.myUI.edt_sql_log_3.clear()  # 清空当前sql日志
        self.myUI.tbw_sql_result_3.tbw_table.setRowCount(0)
        self.myUI.tbw_sql_result_3.tbw_table.setColumnCount(0)

        self.ipr_sql_3 = IndeterminateProgressRing(self.myUI.tbw_sql_result_3)
        self.ipr_sql_3.move((self.myUI.tbw_sql_result_3.width() - self.ipr_sql_3.width()) / 2,
                            (self.myUI.tbw_sql_result_3.height() - self.ipr_sql_3.height()) / 2)
        self.ipr_sql_3.show()

        sql = self.myUI.edt_sql_body_3.textCursor().selectedText().replace("\u2029",
                                                                           "\n") or self.myUI.edt_sql_body_3.toPlainText()  # 获取选中文本
        row = int(self.myUI.edt_sql_row_3.text()) if self.myUI.edt_sql_row_3.text() else 1000

        self.thread_sql_execute_3 = DBQueryThread(self.connection_sql_3,
                                                  sql=sql,
                                                  row_max=row,
                                                  obj=self.myUI.btn_sql_execute_3.objectName())
        self.thread_sql_execute_3.sign_end.connect(self.get_db_result)
        self.thread_sql_execute_3.sign_end.connect(lambda: self.ipr_sql_3.stop())
        self.thread_sql_execute_3.sign_err.connect(self.createErrorInfoBar)
        self.thread_sql_execute_3.sign_err.connect(self.myUI.edt_sql_log_3.setText)
        self.thread_sql_execute_3.sign_err.connect(lambda x: self.myUI.btn_sql_execute_3.setEnabled(True))
        self.thread_sql_execute_3.sign_err.connect(lambda: self.ipr_sql_3.stop())

        # 设置状态
        self.myUI.btn_sql_execute_3.setEnabled(False)
        self.thread_sql_execute_3.start()

    def export_excel(self, widget: QTableWidget):
        """导出Excel数据"""
        if not widget.rowCount():
            self.createInfoInfoBar("表格无数据")
            return
        # 获取文件名
        file_name, _ = QFileDialog.getSaveFileName(None,
                                                   "Save File",
                                                   "data.xlsx",  # 默认文件名
                                                   "Excel 工作簿(.xlsx);;Excel 97-2004 工作簿(.xls);;CSV(逗号分隔)(.csv);;All Files(*)")
        if not file_name:
            self.createInfoInfoBar("取消导出")
            return
        wb = Workbook()  # 创建一个新的工作簿
        ws = wb.active  # 选择默认的工作表
        # 获取表头
        head = [widget.horizontalHeaderItem(col).text() for col in
                range(widget.columnCount())]
        ws.append(head)
        # 获取数据(可以考虑使用子线程后台执行)
        for row in range(widget.rowCount()):
            line = []
            for col in range(widget.columnCount()):
                item = widget.item(row, col)
                line.append(item.text())
            ws.append(line)

        wb.save(file_name)  # 保存
        logger.info(f"导出{file_name}")
        self.createSuccessInfoBar(f"导出成功\n{file_name}")

    def export_mapping(self, widget: QTableWidget):
        """导出mapping数据，数仓专用"""
        if False:
            self.createInfoInfoBar("无mapping表")
            return

        #
        database_name = re.findall(r"[a-zA-Z_]+", self.sender().text())[0]
        item = self.myUI.tbw_table.currentItem()
        table_name = self.myUI.tbw_table.item(item.row(), 1).text()

        # 获取文件名
        file_name, _ = QFileDialog.getSaveFileName(None,
                                                   "Save File",
                                                   "data.xlsx",  # 默认文件名
                                                   "Excel 工作簿(.xlsx);;Excel 97-2004 工作簿(.xls);;CSV(逗号分隔)(.csv);;All Files(*)")
        if not file_name:
            self.createInfoInfoBar("取消导出")
            return

        # tx = self.myUI.tbw_table.currentItem.text()
        # row = self.myUI.tbw_table.currentRow()
        # database = self.myUI.tbw_table.item(row, 0).text()
        # database = "etl_bl"
        # table = self.myUI.tbw_table.item(row, 1).text()
        sql_1 = f"select * from {database_name}.datamapping_task where lower(t_tab_eng_name)=lower('{table_name}')"
        sql_2 = f"select * from {database_name}.datamapping where lower(t_tab_eng_name)=lower('{table_name}') order by cast(seq_num as decimal)"
        connection = self.thread_pool.get_connection()
        cursor = connection.cursor()
        try:
            cursor.execute(sql_1)
            data1 = cursor.fetchall()  # datamapping_task
            cursor.execute(sql_2)
            data2 = cursor.fetchall()  # datamapping
            # os.path.dirname(sys.argv[0]) + "/_internal/aaa_etc/db_config.ini"
            wb = load_workbook(os.path.dirname(sys.argv[0]) + "/_internal/aaa_etc/mapping.xlsx")
            ws = wb.worksheets[0]  # 获取第一个工作表
            ws.delete_rows(idx=2, amount=100)
            ws.append(data1[0])
            ws = wb.worksheets[1]
            ws.delete_rows(idx=4, amount=1000)  # 从第4行开始删除1000行
            for row in data2:
                ws.append(row)
            wb.save(file_name)
            self.createSuccessInfoBar(f"导出成功\n{file_name}")
        except Exception as e:
            logger.error("异常{e}")
            self.createErrorInfoBar(f"异常{e}")
        finally:
            cursor.close()
            connection.close()

    def find_procedures(self):
        """筛选过程清单"""
        # 如果未连接数据库，则不做操作
        if not self.db_info:
            return
        # 获取所有匹配到的行
        database = self.myUI.cbb_find_procedure.currentText()
        find_str = self.myUI.edt_find_procedure.text().strip().upper()
        if database == "所有库":
            find_procedure_list = self.procedures
        elif database == "当前库":
            if self.db_info["type"].lower() == "oracle":
                database = self.db_info["username"]
            elif self.db_info["type"].lower() in ("gbase", "mysql"):
                database = self.db_info["database"]
            else:
                database = self.db_info["database"]
            find_procedure_list = [line for line in self.procedures if line[0].upper() == database.upper()]
        else:
            find_procedure_list = [line for line in self.procedures if line[0].upper() == database.upper()]

        # 筛字符串
        find_procedure_list = [line for line in find_procedure_list if
                               find_str.upper() in (line[0] + line[1] + line[2]).upper()]

        # 加工清单
        self.myUI.tbw_procedure.clearContents()
        # 设置行数
        self.myUI.tbw_procedure.setRowCount(len(find_procedure_list))
        # 向表单添加数据
        for row in range(len(find_procedure_list)):
            for col in range(3):
                item = QTableWidgetItem(str(find_procedure_list[row][col] or ""))
                self.myUI.tbw_procedure.setItem(row, col, item)

    def find_tables(self):
        """筛选表清单"""
        # 如果未连接数据库，则不做操作
        if not self.db_info:
            return
        # 获取所有匹配到的行
        database = self.myUI.cbb_database.currentText()
        find_str = self.myUI.edt_table.text().strip().upper()
        # 筛库
        if database == "所有库":
            find_tab_list = self.tables
        elif database == "当前库":
            if self.db_info["type"].lower() == "oracle":
                database = self.db_info["username"]
            elif self.db_info["type"].lower() in ("gbase", "mysql"):
                database = self.db_info["database"]
            else:
                database = self.db_info["database"]
            find_tab_list = [line for line in self.tables if line[0].upper() == database.upper()]
        else:
            find_tab_list = [line for line in self.tables if line[0].upper() == database.upper()]
        # 筛字符串
        find_tab_list = [line for line in find_tab_list if
                         find_str in (line[0] + line[1] + (line[2] or "")).upper()]
        self.myUI.tbw_table.clearContents()
        self.myUI.tbw_table.setRowCount(len(find_tab_list))
        for row in range(len(find_tab_list)):
            for col in range(3):
                item = QTableWidgetItem(str(find_tab_list[row][col] or ""))
                self.myUI.tbw_table.setItem(row, col, item)

    def get_columns(self, item: QTableWidgetItem):
        """获取选中项目的字段信息 """

        if self.thread_table.isRunning():
            self.thread_table.terminate()
            self.thread_table.wait()
            self.thread_table.connection.close()
            self.thread_table.sign_err.emit("被其他操作中断")

        self.myUI.tbw_column.tbw_table.setRowCount(0)
        self.ipr_column = IndeterminateProgressRing(self.myUI.tbw_column)
        self.ipr_column.move((self.myUI.tbw_column.width() - self.ipr_column.width()) / 2,
                             (self.myUI.tbw_column.height() - self.ipr_column.height()) / 2)
        self.ipr_column.show()

        database_name = self.myUI.tbw_table.item(item.row(), 0).text()
        table_name = self.myUI.tbw_table.item(item.row(), 1).text()
        table_comment = self.myUI.tbw_table.item(item.row(), 2).text()
        self.myUI.lab_table_comment.setText(f"{table_name}（{table_comment}）")  # 设置表显示状态

        file_dict = {"mysql": "_internal/aaa_sql/mysql_get_columns.sql",
                     "oracle": "_internal/aaa_sql/oracle_get_columns.sql",
                     "gbase": "_internal/aaa_sql/gbase_get_columns.sql",
                     "dameng": "_internal/aaa_sql/dameng_get_columns.sql"}

        with open(file_dict[self.db_info["type"].lower()]) as f:
            sql = f.read()

        self.thread_table = DBQueryThread(self.thread_pool.get_connection(),
                                          sql=sql,
                                          sql_parameter={"DATABASE_NAME": database_name, "TABLE_NAME": table_name},
                                          row_max=1000000,
                                          obj=self.myUI.tbw_column.objectName())
        self.thread_table.sign_end.connect(self.get_db_result)
        self.thread_table.sign_end.connect(lambda: self.ipr_column.deleteLater())
        self.thread_table.sign_err.connect(lambda: self.ipr_column.deleteLater())
        self.thread_table.start()

    def get_datamapping(self):
        """查询datamapping"""
        if not self.myUI.tbw_table.currentItem():
            self.createInfoInfoBar("未选中表")
            return
        if self.db_info["type"].lower() not in ('mysql', 'gbase'):
            self.createInfoInfoBar("只能在mysql、gbase数据库中使用该功能")
            return

        database_name = re.findall(r"[a-zA-Z_]+", self.sender().text())[0]
        # return T03_LOAN_CONTR_H
        self.myUI.btn_get_data.setEnabled(False)
        self.myUI.tbw_column.tbw_table.setRowCount(0)
        self.myUI.tbw_column.tbw_table.setColumnCount(0)
        self.ipr_column = IndeterminateProgressRing(self.myUI.tbw_column)
        self.ipr_column.move((self.myUI.tbw_column.width() - self.ipr_column.width()) / 2,
                             (self.myUI.tbw_column.height() - self.ipr_column.height()) / 2)
        self.ipr_column.show()

        item = self.myUI.tbw_table.currentItem()
        # database_name = self.myUI.tbw_table.item(item.row(), 0).text()
        table_name = self.myUI.tbw_table.item(item.row(), 1).text()
        table_comment = self.myUI.tbw_table.item(item.row(), 2).text()
        self.myUI.lab_table_comment.setText(f"{table_name}（{table_comment}）")  # 设置表显示状态

        sql = f"select * from {database_name}.datamapping where lower(t_tab_eng_name)=lower('{table_name}') order by cast(seq_num as decimal)"
        row = 10000
        self.thread_table = DBQueryThread(self.thread_pool.get_connection(),
                                          sql=sql,
                                          row_max=row,
                                          obj=self.myUI.btn_get_data.objectName())
        self.thread_table.sign_end.connect(self.get_db_result)
        self.thread_table.sign_end.connect(lambda: self.ipr_column.deleteLater())
        self.thread_table.sign_err.connect(lambda: self.ipr_column.deleteLater())
        self.thread_table.sign_err.connect(lambda: self.myUI.btn_get_data.setEnabled(True))
        self.thread_table.sign_err.connect(self.createErrorInfoBar)
        self.thread_table.start()

    def get_db_result(self, obj: str, result: dict):
        """查询数据库结果，获取结果"""
        if obj == "get_schemas":  # 获取schema
            schemas = [line[0] for line in result["data"]]
            self.myUI.cbb_database.addItems(schemas)
            self.myUI.cbb_find_procedure.addItems(schemas)

        if obj == "get_tables":  # 获取表清单

            # 加工表清单
            result["data"].sort(key=itemgetter(0, 1), reverse=False)  # 结果集排序
            self.load_tablewidget(self.myUI.tbw_table, None, result["data"])
            self.tables = result["data"]

            # 设置按钮
            self.myUI.btn_connect.setEnabled(True)

        if obj == self.myUI.edt_sta.objectName():  # sta，ods

            self.myUI.btn_yf.setEnabled(True)
            self.myUI.edt_sta.setText(result["data"][0][0].lower())
            self.myUI.edt_ods.setText(result["data"][0][1].lower())

        if obj == "get_procedures":  # 获取过程清单
            result["data"].sort(key=itemgetter(0, 1, 2), reverse=False)  # 结果集排序
            self.procedures = result["data"]
            self.load_tablewidget(self.myUI.tbw_procedure, None, self.procedures)

        if obj == self.myUI.tbw_column.objectName():  # tbw_column

            head = ["表", "表注释", "序号", "字段", "类型", "字段注释"]
            self.myUI.tbw_column.head = head
            self.myUI.tbw_column.data = result["data"]
            self.myUI.tbw_column.setCur_page(1)

            cbb_find_col = [line[3] for line in result["data"]]
            # 清除筛选条件
            self.myUI.cbb_find_col_1.clear()
            self.myUI.cbb_find_col_2.clear()
            self.myUI.cbb_find_col_3.clear()
            # 添加筛选条件
            self.myUI.cbb_find_col_1.addItems(["--条件1--"] + cbb_find_col)
            self.myUI.cbb_find_col_2.addItems(["--条件2--"] + cbb_find_col)
            self.myUI.cbb_find_col_3.addItems(["--条件3--"] + cbb_find_col)

        if obj == self.myUI.btn_get_data.objectName():  # 查询数据
            # 加工表清单
            self.myUI.tbw_column.head = result["head"]
            self.myUI.tbw_column.data = result["data"]
            self.myUI.tbw_column.setCur_page(1)
            # 设置按钮
            self.myUI.btn_get_data.setEnabled(True)

        if obj == self.myUI.tbw_procedure.objectName():  # 获取过程实体
            self.myUI.edt_procedure_body.setText(result["data"][0][0])
            self.myUI.edt_procedure_body.setText(result["data"][0][0])

        if obj == self.myUI.btn_sql_execute_1.objectName():  # sql_body_1

            # 向表格添加数据
            # self.load_tablewidget(self.myUI.tbw_sql_result_1, result["head"], result["data"])
            # self.s = SilenceThread(self.load_tablewidget, self.myUI.tbw_sql_result_1.tbw_table, result["head"], result["data"])
            # self.s.sign_end.connect(self.myUI.tbw_sql_result_1.tbw_table.resizeColumnToContents_new)
            # self.s.start()

            self.myUI.tbw_sql_result_1.head = result["head"]
            self.myUI.tbw_sql_result_1.data = result["data"]
            self.myUI.tbw_sql_result_1.setCur_page(1)

            # 其他设置项
            sql_time = round(result["end_time"].timestamp() - result["begin_time"].timestamp(), 3)
            self.myUI.edt_sql_log_1.setText(f"执行成功！\n"
                                            f"影响行数：{result['affected_rows']}\n"
                                            f"开始时间：{result['begin_time']}\n"
                                            f"结束时间：{result['end_time']}\n"
                                            f"消耗时长：{sql_time}秒")
            self.createSuccessInfoBar(f"影响{result['affected_rows']}行，耗时{sql_time}秒。")
            self.myUI.btn_sql_execute_1.setEnabled(True)

        if obj == self.myUI.btn_sql_execute_2.objectName():  # sql_body_2

            self.myUI.tbw_sql_result_2.head = result["head"]
            self.myUI.tbw_sql_result_2.data = result["data"]
            self.myUI.tbw_sql_result_2.showdata()

            # 其他设置项
            sql_time = round(result["end_time"].timestamp() - result["begin_time"].timestamp(), 3)
            self.myUI.edt_sql_log_2.setText(f"执行成功！\n"
                                            f"影响行数：{result['affected_rows']}\n"
                                            f"开始时间：{result['begin_time']}\n"
                                            f"结束时间：{result['end_time']}\n"
                                            f"消耗时长：{sql_time}秒")
            self.createSuccessInfoBar(f"影响{result['affected_rows']}行，耗时{sql_time}秒。")
            self.myUI.btn_sql_execute_2.setEnabled(True)

        if obj == self.myUI.btn_sql_execute_3.objectName():  # sql_body_3

            self.myUI.tbw_sql_result_3.head = result["head"]
            self.myUI.tbw_sql_result_3.data = result["data"]
            self.myUI.tbw_sql_result_3.showdata()

            # 其他设置项
            sql_time = round(result["end_time"].timestamp() - result["begin_time"].timestamp(), 3)
            self.myUI.edt_sql_log_3.setText(f"执行成功！\n"
                                            f"影响行数：{result['affected_rows']}\n"
                                            f"开始时间：{result['begin_time']}\n"
                                            f"结束时间：{result['end_time']}\n"
                                            f"消耗时长：{sql_time}秒")
            self.createSuccessInfoBar(f"影响{result['affected_rows']}行，耗时{sql_time}秒。")
            self.myUI.btn_sql_execute_3.setEnabled(True)

        if obj == self.myUI.btn_mapping.objectName():
            pass

    def get_gbase_sql(self):
        self.myUI.edt_sta.setText("")
        self.myUI.edt_ods.setText("")

        # 如果没选中表或没有数据库连接，则不操作
        if not self.myUI.tbw_table.currentItem():
            self.createInfoInfoBar("未选择表")
            return
        # 如果没有输入系统名，则不操作
        if not self.myUI.edt_yf.text():
            self.createInfoInfoBar("未输入系统名")
            return
        # 如果链接的不是oracle数据库，则不操作
        if self.db_info["type"].lower() not in ("oracle", "dameng"):
            self.createInfoInfoBar("只能选择oracle数据库")
            return

        # 获取sql
        with open("_internal/aaa_sql/get_gbase_sql.sql", 'rb') as f:
            data = f.read()
            file_encoding = chardet.detect(data)["encoding"]
        with open("_internal/aaa_sql/get_gbase_sql.sql", encoding=file_encoding) as f:
            sql = f.read()
        # 获取sql参数，库名，表名，系统名
        row = self.myUI.tbw_table.currentIndex().row()
        owner = self.myUI.tbw_table.item(row, 0).text()
        table_name = self.myUI.tbw_table.item(row, 1).text()
        system = self.myUI.edt_yf.text()

        self.thread_table = DBQueryThread(self.thread_pool.get_connection(),
                                          sql=sql,
                                          sql_parameter={"OWNER": owner,  # 设置sql参数
                                                         "TABLE_NAME": table_name,
                                                         "SYS": system},
                                          row_max=row,
                                          obj=self.myUI.edt_sta.objectName())
        self.thread_table.sign_end.connect(self.get_db_result)
        self.thread_table.sign_err.connect(lambda: self.myUI.btn_yf.setEnabled(True))

        self.myUI.btn_yf.setEnabled(False)
        self.thread_table.start()

    def get_procedure_body(self):
        """获取选中项目的字段信息"""

        self.ipr_procedure_body = IndeterminateProgressRing(self.myUI.edt_procedure_body)
        self.ipr_procedure_body.move((self.myUI.edt_procedure_body.width() - self.ipr_procedure_body.width()) / 2,
                                     (self.myUI.edt_procedure_body.height() - self.ipr_procedure_body.height()) / 2)
        self.ipr_procedure_body.show()

        file_dict = {"mysql": "_internal/aaa_sql/mysql_get_procedure_body.sql",
                     "oracle": "_internal/aaa_sql/oracle_get_procedure_body.sql",
                     "gbase": "_internal/aaa_sql/gbase_get_procedure_body.sql",
                     "dameng": "_internal/aaa_sql/dameng_get_procedure_body.sql"}
        with open(file_dict[self.db_info["type"].lower()]) as f:
            sql = f.read()
        item = self.myUI.tbw_procedure.currentItem()
        database_name = self.myUI.tbw_procedure.item(item.row(), 0).text()
        procedure_name = self.myUI.tbw_procedure.item(item.row(), 2).text()
        self.thread_procedure = DBQueryThread(self.thread_pool.get_connection(),
                                              sql=sql,
                                              sql_parameter={"DATABASE_NAME": database_name,
                                                             "PROCEDURE_NAME": procedure_name},
                                              obj=self.myUI.tbw_procedure.objectName())

        self.thread_procedure.sign_end.connect(self.get_db_result)
        self.thread_procedure.sign_end.connect(lambda: self.ipr_procedure_body.deleteLater())
        self.thread_procedure.start()
        self.myUI.lab_procedure_name.setText(database_name + "." + procedure_name)

    def get_tables(self):

        self.ipr_table = IndeterminateProgressRing(self.myUI.tbw_table)
        self.ipr_table.move((self.myUI.tbw_table.width() - self.ipr_table.width()) / 2,
                            (self.myUI.tbw_table.height() - self.ipr_table.height()) / 2)
        self.ipr_table.show()

        """获取表清单"""
        file_dict = {"mysql": "_internal/aaa_sql/mysql_get_tables.sql",
                     "oracle": "_internal/aaa_sql/oracle_get_tables.sql",
                     "gbase": "_internal/aaa_sql/gbase_get_tables.sql",
                     "dameng": "_internal/aaa_sql/dameng_get_tables.sql"}

        with open(file_dict[self.db_info["type"].lower()]) as f:
            sql = f.read()
        self.thread_table = DBQueryThread(self.thread_pool.get_connection(),
                                          sql=sql,
                                          row_max=1000000,
                                          obj="get_tables")
        self.thread_table.sign_end.connect(self.get_db_result)
        # self.thread_table.sign_end.connect(lambda: self.ipr_table.stop())
        self.thread_table.sign_end.connect(lambda: self.ipr_table.deleteLater())
        self.thread_table.sign_err.connect(lambda: self.ipr_table.deleteLater())
        self.thread_table.start()

    def get_procedures(self):

        file_dict = {"mysql": "_internal/aaa_sql/mysql_get_procedures.sql",
                     "oracle": "_internal/aaa_sql/oracle_get_procedures.sql",
                     "gbase": "_internal/aaa_sql/gbase_get_procedures.sql",
                     "dameng": "_internal/aaa_sql/dameng_get_procedures.sql"}
        with open(file_dict[self.db_info["type"].lower()]) as f:
            sql = f.read()
        self.thread_procedure = DBQueryThread(self.thread_pool.get_connection(),
                                              sql=sql,
                                              row_max=1000000,
                                              obj="get_procedures")
        self.thread_procedure.sign_end.connect(self.get_db_result)
        self.thread_procedure.start()

    def get_schemas(self):
        """获取所有schema"""
        file_dict = {"mysql": "_internal/aaa_sql/mysql_get_schemas.sql",
                     "oracle": "_internal/aaa_sql/oracle_get_schemas.sql",
                     "gbase": "_internal/aaa_sql/gbase_get_schemas.sql",
                     "dameng": "_internal/aaa_sql/dameng_get_schemas.sql"}
        with open(file_dict[self.db_info["type"].lower()]) as f:
            sql = f.read()
        self.thread_schema = DBQueryThread(self.thread_pool.get_connection(),
                                           sql=sql,
                                           row_max=100000,
                                           obj="get_schemas")
        self.thread_schema.sign_end.connect(self.get_db_result)
        self.thread_schema.start()

    def get_tab_datas(self):
        """查询表数据"""
        if not self.myUI.tbw_table.currentItem():
            self.createInfoInfoBar("未选中表")
            return

        self.myUI.btn_get_data.setEnabled(False)
        self.myUI.tbw_column.tbw_table.setRowCount(0)
        self.myUI.tbw_column.tbw_table.setColumnCount(0)
        self.ipr_column = IndeterminateProgressRing(self.myUI.tbw_column)
        self.ipr_column.move((self.myUI.tbw_column.width() - self.ipr_column.width()) / 2,
                             (self.myUI.tbw_column.height() - self.ipr_column.height()) / 2)
        self.ipr_column.show()

        item = self.myUI.tbw_table.currentItem()
        database_name = self.myUI.tbw_table.item(item.row(), 0).text()
        table_name = self.myUI.tbw_table.item(item.row(), 1).text()
        table_comment = self.myUI.tbw_table.item(item.row(), 2).text()
        self.myUI.lab_table_comment.setText(f"{table_name}（{table_comment}）")  # 设置表显示状态

        sql = f"select * from {database_name}.{table_name} where 1=1"
        # 获取条件
        if self.myUI.cbb_find_col_1.currentIndex() > 0:
            # print(self.myUI.cbx_is_sensitive.isChecked())
            # sql = (sql +
            #        " and upper(" +
            #        self.myUI.cbb_find_col_1.currentText() +
            #        ") like upper('%" +
            #        self.myUI.edt_find_str_1.text().strip() +
            #        "%')")
            # sql = f"{sql} and upper({self.myUI.cbb_find_col_1.currentText()}) like upper('{self.myUI.edt_find_str_1.text().strip()}')"
            sql = f"{sql} and {self.myUI.cbb_find_col_1.currentText()} like '{self.myUI.edt_find_str_1.text().strip()}'" if self.myUI.cbx_is_sensitive.isChecked() else f"{sql} and upper({self.myUI.cbb_find_col_1.currentText()}) like upper('{self.myUI.edt_find_str_1.text().strip()}')"
        if self.myUI.cbb_find_col_2.currentIndex() > 0:
            # sql = f"{sql} and upper({self.myUI.cbb_find_col_2.currentText()}) like upper('{self.myUI.edt_find_str_2.text().strip()}')"
            sql = f"{sql} and {self.myUI.cbb_find_col_2.currentText()} like '{self.myUI.edt_find_str_2.text().strip()}'" if self.myUI.cbx_is_sensitive.isChecked() else f"{sql} and upper({self.myUI.cbb_find_col_2.currentText()}) like upper('{self.myUI.edt_find_str_2.text().strip()}')"
        if self.myUI.cbb_find_col_3.currentIndex() > 0:
            # sql = f"{sql} and upper({self.myUI.cbb_find_col_3.currentText()}) like upper('{self.myUI.edt_find_str_3.text().strip()}')"
            sql = f"{sql} and {self.myUI.cbb_find_col_3.currentText()} like '{self.myUI.edt_find_str_3.text().strip()}'" if self.myUI.cbx_is_sensitive.isChecked() else f"{sql} and upper({self.myUI.cbb_find_col_3.currentText()}) like upper('{self.myUI.edt_find_str_3.text().strip()}')"
        row = int(self.myUI.edt_result_row.text()) if self.myUI.edt_result_row.text() else 1000

        self.thread_table = DBQueryThread(self.thread_pool.get_connection(),
                                          sql=sql,
                                          row_max=row,
                                          obj=self.myUI.btn_get_data.objectName())
        self.thread_table.sign_end.connect(self.get_db_result)
        self.thread_table.sign_end.connect(lambda: self.ipr_column.deleteLater())
        self.thread_table.sign_err.connect(lambda: self.ipr_column.deleteLater())
        self.thread_table.sign_err.connect(lambda: self.myUI.btn_get_data.setEnabled(True))
        self.thread_table.start()

    def init_connection(self):
        """连接成功时，初始化部分数据库连接"""
        self.connection_sql_1 = self.thread_pool.get_connection()  # sql窗口1的连接
        self.connection_sql_2 = self.thread_pool.get_connection()  # sql窗口2的连接
        self.connection_sql_3 = self.thread_pool.get_connection()  # sql窗口3的连接

    def init_ui(self):

        # 初始化标题标签页
        # 删除部分组件
        self.myUI.pushButton_1.deleteLater()
        self.myUI.pushButton_2.deleteLater()
        self.myUI.pushButton_3.deleteLater()
        self.myUI.pushButton_4.deleteLater()
        self.myUI.pushButton_5.deleteLater()
        # 添加部件
        self.myUI.segw_1.addItem("tab_table", "表预览", onClick=None, icon=FluentIcon.CALENDAR)
        self.myUI.segw_1.addItem("tab_proc", "过程预览", onClick=None, icon=FluentIcon.LAYOUT)  # LAYOUT
        self.myUI.segw_1.addItem("tab_sql", "SQL执行窗口", onClick=None, icon=FluentIcon.LABEL)
        self.myUI.segw_1.addItem("tab_db", "数据库配置", onClick=None, icon=FluentIcon.SETTING)
        self.myUI.segw_1.addItem("tab_settings", "系统设置", onClick=None, icon=FluentIcon.DEVELOPER_TOOLS)
        self.myUI.segw_1.setCurrentItem(self.myUI.tab_table.objectName())
        self.myUI.segw_1.currentItemChanged.connect(
            lambda k: self.myUI.stackedWidget.setCurrentWidget(self.findChild(QWidget, k)))
        # 初始化表预览页
        self.myUI.tableWidget.deleteLater()
        self.myUI.tbw_column.head = ["表名", "表注释", "序号", "字段名", "类型", "字段注释"]
        self.myUI.tbw_column.setCur_page(1)

        # 初始化sql标签页
        # 删除部件
        self.myUI.pushButton.deleteLater()
        self.myUI.pushButton_9.deleteLater()
        self.myUI.pushButton_10.deleteLater()
        self.myUI.tableWidget_2.deleteLater()
        self.myUI.tableWidget_3.deleteLater()
        self.myUI.tableWidget_4.deleteLater()
        # 添加部件
        self.myUI.segw_sql.addItem("tab_sql_1", "SQL执行窗口一")
        self.myUI.segw_sql.addItem("tab_sql_2", "SQL执行窗口二")
        self.myUI.segw_sql.addItem("tab_sql_3", "SQL执行窗口三")
        self.myUI.segw_sql.setCurrentItem("tab_sql_1")
        self.myUI.segw_sql.currentItemChanged.connect(
            lambda k: self.myUI.staw_sql.setCurrentWidget(self.findChild(QWidget, k)))
        # 初始化sql_1执行结果标签页
        # 删除部件
        self.myUI.pushButton_11.deleteLater()
        self.myUI.pushButton_12.deleteLater()
        # 添加部件
        self.myUI.segw_sql_1.addItem("tab_sql_1_result", "执行结果")
        self.myUI.segw_sql_1.addItem("tab_sql_1_log", "执行日志")
        self.myUI.segw_sql_1.setCurrentItem("tab_sql_1_result")
        self.myUI.segw_sql_1.currentItemChanged.connect(
            lambda k: self.myUI.staw_sql_1.setCurrentWidget(self.findChild(QWidget, k)))
        # 初始化sql_2执行结果标签页
        # 删除部件
        self.myUI.pushButton_16.deleteLater()
        self.myUI.pushButton_17.deleteLater()
        # 添加部件
        self.myUI.segw_sql_2.addItem("tab_sql_2_result", "执行结果")
        self.myUI.segw_sql_2.addItem("tab_sql_2_log", "执行日志")
        self.myUI.segw_sql_2.setCurrentItem("tab_sql_2_result")
        self.myUI.segw_sql_2.currentItemChanged.connect(
            lambda k: self.myUI.staw_sql_2.setCurrentWidget(self.findChild(QWidget, k)))
        # 初始化sql_3执行结果标签页
        # 删除部件
        self.myUI.pushButton_19.deleteLater()
        self.myUI.pushButton_20.deleteLater()
        # 添加部件
        self.myUI.segw_sql_3.addItem("tab_sql_3_result", "执行结果")
        self.myUI.segw_sql_3.addItem("tab_sql_3_log", "执行日志")
        self.myUI.segw_sql_3.setCurrentItem("tab_sql_3_result")
        self.myUI.segw_sql_3.currentItemChanged.connect(

            lambda k: self.myUI.staw_sql_3.setCurrentWidget(self.findChild(QWidget, k)))

        # 设置窗口图标
        # 图标文件(二进制)
        # binary_data = b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00d\x00\x00\x00d\x08\x06\x00\x00\x00p\xe2\x95T\x00\x00\x00\x01sRGB\x00\xae\xce\x1c\xe9\x00\x00\x00\x04gAMA\x00\x00\xb1\x8f\x0b\xfca\x05\x00\x00\x00\tpHYs\x00\x00\x0e\xc3\x00\x00\x0e\xc3\x01\xc7o\xa8d\x00\x00\x17\xdcIDATx^\xed\xdd\x07\xb0%E\xd5\x07\xf0\x05\x94\x9c\x93\x08H\xceA\x92D\x95$ B\x91\x8bX\x05\x88\x92\x93H\x06\x15K%\xa3d\x04\x15\x94\x9c3\x14J\x96(\x14 9\'A\x10%g\t\xe2|\xdf\xaf\xab\xfe\xaf\x9a\xcb\xdd}\xf7\xee>v\xdf]\xdf\xa9:5=3==\xdd\xe7\xdf\xe7\xf4\xe9\xd3=\xf7\x0ek\x86hP\xd1\x10 \x83\x8c\x86\x00\x19d4\x04\xc8 \xa3!@\x06\x19\r\xfb\xe4\x93O\x9a\xff\xfe\xf7\xbf\x85kr\xfd\xe3\x8f?n\xfe\xf3\x9f\xff\x94#J\xda1\xcf9:\xef\x8f\xeb\xf7\xf4\n\xab\xb7\xf6J\xa7\xfe\xce\xc3\xb5\x0c\xda\xa5=\xdf-}\x06\x90\xba\xd0p\n\xae\xcf?\xfa\xe8\xa3>pp\x9d\xbf\x1d\xe7\x1d\xbd\xc2H\xbb"\xfc\x0f?\xfc\xb0\xb4\xa3nk\xf2F\x06\xd2\xf5=\xd7\xba\xa5a)4\xa4\x10\xc2\xce\xf1\xed\xb7\xdfn^\x7f\xfd\xf5\xe6\x95W^)\xfc\xea\xab\xaf\x16~\xf9\xe5\x97\xcb\xb9{o\xbc\xf1F\xf3\xdak\xaf\x8du\xacmX{\x1d\xdf|\xf3\xcd\x92\xfe\xd7\xbf\xfeU\xda\xfe\xce;\xef\xf4\x81\x81#\xcb\x00\xe2\xd8-\x951\xa4F[oH\xefW\xa9\xdbo\xbf\xbd\xf9\xf5\xaf\x7f\xdd\x1cq\xc4\x11\xcd/\x7f\xf9\xcb\xe6\xc8#\x8fl~\xf5\xab_\x95\xa3\xf3\xa4\xc76\xd6\xae\xa3\x8e:\xaa\xb0\xf4\xd1G\x1f\xdd\x1cs\xcc1\xe5\xdea\x87\x1d\xd6\x9ct\xd2I\xcd_\xff\xfa\xd7\xe6\xddw\xdf-\xf2\x03H\x00\x88,q\xb7TLV\n\xab\xd3\x00y\xee\xb9\xe7\x9a\xe3\x8e;\xaeY`\x81\x05\x9a/\x7f\xf9\xcb\xcd\x8c3\xceX\x8e3\xcc0CI\x7f\xe9K_j\xa6\x99f\x9af\xba\xe9\xa6+\xd7\xc66N[\xa7\x9dv\xda\xbe6\xe6\xda\xb2\xcb.\xdb\x9cu\xd6YES\x08>\xf2\x1bU\x1aF\x1bR\x18\xaeU\xef\xd9g\x9fm\x0e<\xf0\xc0R\x89\xf1\xc6\x1b\xaf\x19g\x9cq\x9aq\xc7\x1d\xb7\x1c\xbf\xf8\xc5/6_\xf8\xc2\x17\xcau\xd7\x86\r\x1b6B\xf6L/\xb16i\x9fv\xd6m\x97\xd6\x9e\xf9\xe6\x9b\xafh\xc9?\xff\xf9\xcfO\x01\x12\xd9\xe1\x91\xa1a4!\x05\xc4\\\x05\x14\x1ar\xf0\xc1\x07\x17mP\xa1\xf1\xc7\x1f\xbfh\xc5\xac\xb3\xce\xda\xcc2\xcb,\xcd\xcc3\xcf\xdc\xcc4\xd3L\x85\xa5G\xc4_\xf9\xcaWz\x8a\xd3&im%\x83\xa9\xa7\x9e\xba\x80D\x16\x0b.\xb8`s\xca)\xa7\x94\xb14` rK\xc7\x1e\x19*\x83:R\x00@jP\x9e\x7f\xfe\xf9\x02\x08\x15\x05\xc6\xbc\xf3\xce\xdb\xec\xb3\xcf>\xcd\xb9\xe7\x9e\xdb\\z\xe9\xa5\xcd%\x97\\\xd2\\t\xd1E\x1d\xf1\xc5\x17_\\\xf2\xf7\n\xa7\xbe\xda\x89\xcf<\xf3\xccf\x97]v)\xc0L0\xc1\x04EC~\xff\xfb\xdf\x17\x93EV\x01\x04\x91i:u\xb7\xd4\x07\x08\xb7.\xda\x12p\x00r\xe8\xa1\x87\x96\xdeBU\x8d%*\xc1\xcb\xf8\xf7\xbf\xff\xdd\xbc\xf7\xde{\xc5\xd30\xb0I\xf7\xc7\xef\xbf\xff~\xcf\xf0\x07\x1f|P\xda\xa8}\xea\xce4\x1d{\xec\xb1\xcdTSMUL\xd6B\x0b-TL\x16YD#\xc2(r\xec\x96\xca\xa0\x8ej4\xa51@\x0e9\xe4\x90\xa2!l\xe9\xc2\x0b/\\\x062\xae\x1f-RiGy\x95#\x1d-S\x99\xe4\x91N\x99\xc9\x87\x93Ggp\xcfs\xee\xe7yi\x9c\xe7\xeak\xa3\x83\xbd/\xef$x\xde&\x93m,\xa1!\xbf\xfd\xedo?e\xb2RO\xc7\xc8\xb1[\xea\x03\xa4&\x05\xe1\x1a\x90h\x08\xd55\xef\xc8\xcbU\xe8\xdak\xaf\xeds\x07\xdd\x7f\xea\xa9\xa7\xca}\xf9\xfe\xf4\xa7?\x15\xf7\xf8\xa0\x83\x0ej\xce>\xfb\xec\xe6\x89\'\x9e(\x8d\xd4\xf3\xee\xbc\xf3\xcer\xefg?\xfbY\xe9m\x0f=\xf4P\xe9\x95u\x03\xc3\x9e\x19\x9d\xe4}\xe9 H;\x012\xfd\xf4\xd3\x17@\x8c!\xbf\xfb\xdd\xef\xca\xf5\x80\x86R\xdf\xa4\xbb\xa5\x8e\x00Q\t\x83\xd9\xa2\x8b.Z\x84J\xd0\xe9\xe1\x8f?\xfex\xf3\xf3\x9f\xff\xbcY|\xf1\xc5\x8b\x1ao\xb1\xc5\x16\xcd\xad\xb7\xdeZ\x9e\xe7\x14\xf0\xd2\x96^z\xe9\xd2\xa3\xbe\xfb\xdd\xef67\xdcpC\xb9\xa7!\xbf\xf9\xcdo\x8a\xd6\xcd9\xe7\x9c\xcd\x06\x1bl\xd0\\q\xc5\x15\xa5l\x14P\x92\x0e \xa9\xdb\xe8d\x14@\xb8\xc0\x06u\x9ds\x8c\x01BM\x01\xb2\xc8"\x8b4g\x9cqF\x1f \x9e}\xf0\xc1\x07\x9b\xddv\xdb\xadh\xd1\xe4\x93O\xde\xac\xba\xea\xaa\xcdu\xd7]Wz\xd7\xd3O?\xdd\xec\xb9\xe7\x9e\xc5[\x99x\xe2\x89\x9b\xef|\xe7;\xcd\x95W^Y\x9ec\x02L\xbax.\xb4o\xc9%\x97\xec3\x87\xde-\x0f\x96\xd6`\x9cz\x8d.\xf6N\xed\xc01Yc\\C\x0c\xea<\x0b\x80|\xf5\xab_mN?\xfd\xf4\x02H\x04\xf6\xe4\x93O\x96<\xdf\xfc\xe67\x9b\xaf}\xedk\xcd\xf7\xbf\xff\xfd\xe6\x8e;\xee(\x8dx\xe1\x85\x17\x9a\xc3\x0f?\xbc\xf9\xc67\xbeQ4a\xeb\xad\xb7n\xfe\xfc\xe7?\x97\xe7\x08\x9e\x83\xb0\xcc2\xcb\x94\xde\xb6\xe1\x86\x1b6\x7f\xfc\xe3\x1fKx\xc2\xbb\xe5I=\xf2\xae\x9c\x8f.n\x05\x84Y\r \xf3\xcf?\xff\x98\xd3\x10\xbd\x1f K,\xb1D\xe9\xc55 \x04{\xcb-\xb7\x94\x01\xee\x84\x13N(\xee" h\x908\xd8\xcd7\xdf\\\xee\x99\xf1s\x1f\x9fy\xe6\x99\xd2\xc0\x8c!\x1a),q\xfe\xf9\xe7\x17p\x8d!i nM\xa7n\x9f7\x07\x8c\xbc\x93{\x1b@\x98,&x\x8c\x00b\x1e\x921\x84\x06\x00D\xa0-\x02RiB$|\xb1/=\x1c\x18\xbc+y\xb8\x8c\xee\xbd\xf5\xd6[%\x1d\xf7\x1as/]\xf7\x9c<\xces/e\xcb\x9ft\xcc\xe4\xe8`\xef\xca\xfb\xc8\x82\x86\x9cx\xe2\x89\xc5|3\xb1cLC2\x0f\xe1\xf62Yz\xbb\xc1\x1a(\x89\x88\xd2\x18\xc7:Ms\x1c\xf5,iy\x93?\xec^k\xfe\xb0\xeb\xf2\xbb.-/\xb0\xeb<\xa3\x8b\xd5\x83\xf3"\xc8(\xa6\xa5s\x8eQ\x93\x15\xdf[\xc8d\xf7\xddw/\xae-M1\xc0\xb7\xb2{\xa7\x9dvZ\x19k\x92\x87gv\xce9\xe7|\xea\x19\xf7O=\xf5\xd4\x92\xdf\xbd\\s\x94\xcf\xf5\xfa\xfc\x0f\x7f\xf8C\xc9\xe7z\xa7\x9cr\x92n\xe5:o\x9d\'\xe7\xa9\x93\xf4\xc9\'\x9f\xdcl\xb7\xddv\xcd$\x93L\xd2\xd1\x18\x12\xee\x96:\x1eC\xd8M\xaa\xca\xed3\xc8cAG`\r&f^[#\xb3a\xe7a\xf5w-\xf9=\x9b<u\xbaf\x91m!\xa4\xcc\xd4\x13:\x89\x89\r\x00d\xca\xdc\x8d\x0cu\xac!\xd4TE\x80\xa2\x87\x00\xc81\xe9\xc1\xc2\xea7\xe1\x84\x13\x96\xfa\xb6\xbb\xdfZg\xf9\x13\xb9\xae\xf3\xb4\xcb\x8b\xe5\x07\x8a)\x80yT\x1d\xed\xad\x81q\x1e\x80\xba\xa1\x8e\x00\xd1\x9b\x02\x84X\x0e-q\x9cr\xca)\xfb\x8e\x83\x85\xa7\x98b\x8af\xb2\xc9&+is\x9c\xb0\xde\x8d\xdb\xd5\xd7\xfc\xc93\x93N:iy>\xf9\xe5\xc5\xd18G\xe7\xf2s\xe3\x994\x1a\x12y\x01\x04G\x86\xc0\xe9\x96\xfa\x05D\xc8C%\x00bF\xbd\xeb\xae\xbb\x16\xdb\xc9\xae\xb2\xafa\xe3\xc1\x88\xd8\xb82:\x99\xb0\xb0w\x0b\x933/\xad\xf5\xa9\xf3\x18\xa3\xf2\xact8\xf9\x1c\xe5K[/\xbc\xf0\xc2\x12\xea\xe19f\x0c\t\x93\x1dj\'\xdb\xfe\xa8\xe31\x84ZSS^\x96\x85+\xde\x8f\xde\x11\xcf)^\xd4`a\x83-WU\x1d\xd5\x15\xbbVs}\x9d\xe9\xa9\xf3+#\xf7\xa5]\xcf=G\x1e\x1f7\x9d\xfcb\xaaj\xd9%\xdd-u5\x86,\xb6\xd8b\xc5\x0b\xe1\x86zN%\xd2;\xf2\xcc\x98fu1\xa0\xc6|`\x82\xfb\xc7?\xfe\xd1\xdc}\xf7\xdd\xcde\x97]Vz\xb8I\x9ep\xbau\xf2\xe3\x8f?\xbet\xb4\xf3\xce;\xafLr\x05G\xcd\x8b2\x17\xa9\xdbY\xb3\xf7\xa1\xdc\xaf\xeb1\xb2\xd4/ \x99\x18\x1a\xc8\x04\x10\x01\xa2\xc7\xa4\x12\xa9h]\x991\xc9\xea\x92\x1e\xab\x9ew\xdduW19?\xfd\xe9OKp\xf3\xdb\xdf\xfev\tv\x1a\x03,\xb8aiQ\x88\xe5\x97_\xbe\xd9d\x93MJ\xfcM\xdcJ\x14\x9b5\xc8\xf2\x00V.\xf2\x1e\xe9\xfa\xbda\xe7#K\x1d\x9b,\xabd4\xc4\xd8\xa1\xa1\x1a\x9d\x97\x0f&&$\x1a|\xcf=\xf7\x94^\xbf\xcd6\xdb\x14\x008&\x06nAN^\x9811\x1e\x96\xb6\xe9p\x8e\x06l\x1d\xd0<c\xddu\xd7-\x1dR\xb0\x94,\xa2y\xde\xa3\xed\xed\x00\xc9y\xb8[\xea\xcad%\xfc\x9e\x00`\xfd\xc2\x9c\x0f\x8f\x07\x8aRVk\xf9XO\xd6\xa3\xc5\xc5\x042\x05-yL\x01\x80\xd0yU\xc0\x99}\xf6\xd9\x8bv\x10\xfc\xdcs\xcf]"\xd2<+\x13\xbf\x00%m2\xbc\xd6Zk\x95mP\xb6\xfd\x88\xc1\x91\x19p";\xef\x0e \xe1\\\xeb\x96:\x02\xc4\xa4H%G\x04\xc8\xe8\xa2\x98\x8cZC]#(K\x01\xc2\x1b_\xff\xfa\xd7\x8b6\xe8Dz=w\xd6F\x85\xa5\x96Z\xaa\xac\xbb\xfc\xe0\x07?(\xde\xa3\xf1\x83i\xf2\xcc~\xfb\xed\xd7l\xb9\xe5\x96\xcd\x8a+\xae\xd8\xcc3\xcf<\xc5\xc5\x05$\x10\x95\x05\xb4\xef}\xef{\xcd5\xd7\\S\xe2o\xc3\x1b[\xeak\xea\xd8-\xf5\x1c \xde\xa9\xce\xe9\xa1\xce\xad\xe9\xdfv\xdbme]\x860\x01\xa0\x873?\xb4a\xf5\xd5W/B\xb78&\xa2\xcc\xa3\xe2\x19rY\xb5\xe5\xa5\x97^*\xde\x13SG\x0b\xb8\xbb[m\xb5Uio\xb4F\x99\xb4\xcb\x18\xc41\xf0\x1c\xc1\x0b\xa2f\xcc\x8a,G\x16\x0c\xd4\x93\x80\xd0\x86\x08@\xa4Y\x18\xff\x87?\xfca3\xdbl\xb3\x95\xc9+\xad\xd0\xbbi\x84\xd9\xf4c\x8f=V\x04Nx\xa9\xb3#P\x13}v\x9e\xf2h\xc0\xdf\xff\xfe\xf7"x\xc0\xd8\nD\xdb\x80\x0c\x98\x15VX\xa1\xecH\x01`\x9eK}\x9cK+{d\xa8\'MV\x1a\xad\xee&g{\xec\xb1G3\xc7\x1cs\x14 \x08\x8e\x13\xb2\xd3N;\x95\x81\x1dx\xd9EBHuoNY\xc8y\xc0\x91v\xa4y\x8f>\xfah1mv**_\xf8\x84\xe6\xb5j\x8agR\xa6\xe3\xc8\xca\xa6\xe7\x00\x89) `\xcb\x00V$\xb9\xe3\x13M4Qa\x039\x17\x97 \x01\x91\x05\xaf\xd8\xfcvu\x8e\x0c\xdcS~\x00\xc4\xe6#\x16\xd5\xb8\xfbk\xac\xb1F\x01\x85,D/6\xddt\xd3\xe2\x81\xc9\xe3\xd9\x1a\x88\x94\xd5-\xf5\x1c \xeaK\xb8\x84`"\xb7\xca*\xab\x94A\xd7\xe0k\x15\x0f\x18L\x94<\xea\x07\x90\x08J:\x009w=\xed\xaf\xcf[{\xbcg\xac\x82j\xfb\xca+\xaf\\\xb4\xc4\xfb\x98\xc8\x1f\xfd\xe8Ge\xad$\xcf\x84\x94\xe7Z\xb7\xd4\xb3\x83\xfa}\xf7\xdd\xd7\xec\xb0\xc3\x0ee\xf1\xcc\xb8\xe1\xe8\xdcl\\\x0f\xd7;\x81\x82\xf5tB\xadM\x12N\x1b\xeas\xc7<\'/&Xe\x18Wxe\\e\x80\xd0\xc8o}\xeb[\xa5c\x08\xbf\xc8\x972\xc3\xddRO\x02B;,\x18YRfB\x98\x8f\xb5\xd7^\xbb\xac\xe7\xf3\x9c\xb4)\x82\x94\x1f\x108\x82Oo\x0e\xa7-8\xf7\x03F\xce\x01\xa4\x0c\xfb\xcal\xa7\xe5\x16g\xbc\xe2F\xdf\x7f\xff\xfdm\xcb\xeb\x96z\n\x90\xbc\xf3\x91G\x1ei\xb6\xddv\xdb2aU/[rl\xd2S\xdfP]\xb7\x08)\xe9\x9c\'\x1dJ\xf9I\xd7\xf9\x00\xe2HS\xec\x9c\xb1\xdd\xc9\xac\x9f\xe7e\xeeb2\x9a\xb1\x04\xd7\xda\xd2\r\xf5\x0c y\x9f^kC\x9d\xc9\x1f\x81\xf0x\xd6[o\xbd\xe6\xea\xab\xaf\xee\xeb\xd1\x03E\xca\xc2\xde[\x1fE\x85m\x0e4\xd9\xa4%<<c\x979\x0e\x92\xaf\x06\x04wJ=\x05\x88\xba\x9a\xd0\x99\xe4\x11\x82:\x99A\x13\x8e\x08\xad\xfbz\xf2@P\xde\x17\xf9\xa4\xbd\xce\x99\xae\xcb/\xbf\xbc\x04#u\n\x8bZ\x9bm\xb6Y\xd9\xf2T\xe7\xcf1\xe9N\xa8\xa7\x00\xa1\x01l\xf8\xf6\xdbo_f\xd0z\'\xa1\\p\xc1\x05\x9f\xfa\xb4l\xa0Hy\xd8\xbb\x93\x0e\xe01\x9b\x02\x91d\xb3\xdcr\xcb\x95M\x18L\x1a\x8a|\xc2\x9dRO\x01B\x18\xbey\\s\xcd5\xfb\x96^7\xdex\xe3\xd23\x815\x90` \xef$\x9f\x94-M;\\\x17n\x11p\xb4\x8a\xca\r\xe6y\xd9\xf0\'\x12._\xea\xd2\xad\x9cz\n\x10\xc2\xb0\xdd\x94w\xc5\xe54\xa8\x0b\x99\x98\xad\xd7B\x18\x08J\xfb\xb0r\xd3\xd6\x1ci$\x8d\xe0P\x00\xc4\x9c\xe4\xc7?\xfeq\x99D\xaaK4)etJ=\x05\x08s\xc0\xb5\xd5\x1by7B\xe3v\xd7\x13\x82\xfb\xda2PuRVd\x93\xb6\xe2\xba\xe7_\x7f\xfd\xf5e\xad\x85\xe9\x14\xbe\xe7\xfe>\xf0\xc0\x03\xe5\x1e\xad\xaa\x9f\xeb\x94z\x0e\x10\x9f\xd3\x19\xc8\x01\xe2\xc8\xdd\x05\x08A\xd5=yTIY\x018ecZ\xaa\xf7\xbbn\xb9\xd7\xcc\x9d\xb6\x92\x91\xb1M\xfc,\xcf\xd4\xa0tJ=\x05\x88\xd963\xc1n\xeb\x95\x00Q?\x9f=\xa4G\x0fT\x9dR\x8e2\t7\xed\x05F\xc6\x91\x1bo\xbc\xb1\xccG\x84\xe7\x99O\x83\xbcHA\xea\x11\xe0p\xa7\xd43\x80 \x82`\xb2\x00!t1\xd7\\s\x95\xaf\xafx^\xea\x12PFD\x9d\xd69\x03\xb9\xfc\xd1\x8e\x94OS]\xb7X\xe5S\x0b\xd1\x02\xeb.>\n\xcd\x8c=\xf9\xbb\xa5\x9e\x03\x84\xff\xaf\x1e\xd6;L\xcc\xf6\xddw\xdf\xe2\x82\xa2\xfe\xea\x94\xfb\x9d\xd4\x9b\\\xf4p\x1c@\xa4s\xdd\xb95\x11\xfb\x0c\x98O\x83\xba\x90\x8a\xc0\xe6\xff\x0c \x04a ]m\xb5\xd5\x8a\x990\x17\xb1\x89\xc1\xce\x12\xed\xe8O\x00\xa9s\xa7\xf5\xd6\x01\xea\xe80\xad\xc9{,b\xd9Jd\x85\x92l\xac\xcf\xfbR\xc0v\xa3\xe4\xe9\xaf>\xed\xa8\xa7\x00!\x18&\xc1v\x1e\x01E\xa6\xc2grW]uU\x11\\\x06\xd1\x11Q\xa7\xf5\x96Gy\x01\xc1y\x84,\xcd\x91\xd8k\xaf\xbd\xfa>\x19\xb7&c\xe97\x00\xea<#C=\x05\x88\xbaZ\x0f\xe7\xef\xb3\xd9\xfc\x7f{\xaa\x84\xc4\xc5\x97"\xb0\x81\xa0\x98\x1cmL\xb9\xc0\xa15\x8e&\xa8:\x83\x8e!|b\xf1\xca\x1cI\xfe\x0022\xf2\xe9)@\x90\x86\xdag\xeb\xe3!\x03\xbb\xd0\x05w\x93\xd9J\x8f\x1e\x88z\x05\x04\x1cpb\xc2Dui\x03o/\x03zB\xf0\xf2F\xab\xa4\xbb\xa5\x9e\x03\xc4;}T\xba\xce:\xeb\x94\xa0\x1eP|8jC\xb5M\x07\xc93\xaau\xf3|\r\x88\xf3\xf4|N\x84\xcf\xbf\xbd\x9f\x96\xea\x1c>w\xb3HUw\x8a\x911[=\x05\x88\xf7q9_|\xf1\xc52C7\xa0\x9a\x8f\xd0\x12{\xa6\xfe\xf2\x97\xbf\xf4\xd5kT\xeb\x160Z\xd9Fk1+Q\x02\xdaa\xf9\xd8\xf6S\xf1\xb4\x00\'\x9f\xe3\xff\x84\x86\xc4\x1cX$b\xc3\tD\xdd\xac\xa7\x0b\xc3\xff\xedo\x7f\x1b\x90\xbay> \x84\xc5\xaf\xec\xf7\xb5 \xc5\xed\xa6\x1d\xdekW\x8a5wy\x80\x80\x99\xb7\x91\xa1\x9e4YLAz\xaaO\xcb\xd4M\xf8B\xd0\xd15\x91\xd8\x90\xf6\xc9\x9fvz>f%T\xdf#\xcc\x9c;\x86m\'2\x90\xf3\xf0\x0c\xe24S\xc4\x99\xdbmv^{y\xf2\'\x1d\xee\x94zrPWg=\x90 \xec\xbf\x8a\xebI@f\xce\xbc.u\xcf\x8c\x1aG\xb3\xa4\x1d1\xe1\xbb\x1e\xd3\x92{iW\xf2\x18\xc4\xc5\xad|\xf4)Db\xdcb\xae\xc4\xb1L\x0e-\x9a)G\x9d\xd4\x0f\xe7\xbd\xe1N\xa9\'M\x96\xf7FP\\M\xfb\xa3l:0c\xb6\xb9\xda\xa66\xf56k\x0e(\xf2GX\xd2\xda-\x1d\xa0r?\xe5\xe7>Sd\x01\xcc n\xdc \x07\xbb\\\xec\xff\xf2c\x08\x01>\xe5be\xa4\xdcn\xe5\xd4\x93&+\r\'\x04\xf3\x0f\x1b\x0c\xd6_\x7f\xfd2\xb83%f\xf1\xe6\'v4\xda]h\xfb\x8e\xde[\x0b\xabUh)3\xe0\xe8\xf5\xbc9\x8bP\xe6\x18\xb4P\xd9\xb4\x03\x18\xbf\xf8\xc5/\x9a\x87\x1f~\xf83[\x8eR\x86\xb4\xeb\xdd\xca\xa9\xe7\x00Az\xa4\x86G\x08\xbc.{\xa3\xac\x1e2)z0\xc1\x89/\x19\xf8\t\x8fi\xb1\x97Kh\xc3x\x10\xe1i\x83#\xc1Z\xed\xa3Uv#rcyn>\xe3\xe3\xde2Q\xf6\xf5\x92\x01\xe7\xc1\xceH[\x8e<C\x1e\xd1\x92\x80\xa0\xcc\x94\xdf\x8d\x9cz\x12\x10\x03\xa8zc\x02\xd0\xfb\r\xe4\x02\x8f\x84h\xc2f\x90\xe7\t925\xec\xbd\x8dtz\xbc6\x08\xb7\xd8\r\xcf]\x15\xb5e\x96l\xcc\x16 \xf4\xa1\x8e\x95@\xe6\x0f\x10\x99\x80\x02\x97\xd3@3\x08\x9f9\xa3\x81~\x13L\xa7 \x8f\x00QkG7r\xea9@\xea\xc6&\x1dPbfl\xc91\xb8\xfb|\xdb\xb8\xc2\xd4\xf0\x8c\x0c\xfaL\x0fW\xd5\xcex\xdbP\x99#\xbb\xd9M\xee\xecd\xa1aL\x9eg<\xeb\x19\xe0\xf0\xae\xfcx\x0e\xef\xce\xfb\xbcK\xc4\xc0\xc7<v\xc8\xd3\xaa:\x8eE\xe3\xa4\xbb\x95S\xcf\x01B\xf0\xea\xec\xdd\x04\xa0\xf1\xd2\x8e\xce\x99\x0e\xbdU\xaf\xf5\x15\x95:\xab\xbf\xf9J>\xc0!h\xc7h\x91\xb4\xf6\xe5\x87\x11\x80\'\x92,\x82+"\xe0\xd7\xf2\x98;\x11^`h\xbf\x1f\xf6\xa41\xcc\x19\xd7\xdb\xca\xa5\x852 \xa8\x9f#\xeaVN\x1d\x01b\xbb\xa4\x8a\xb2\xa7\x83\xc1d\xc5\xbd\xf4\xfe4\xde\x11\x13\x98\xa3\x1e\xaa\xfe\xcc\xd1\x01\x07\x1cP\x04k0fz\xb2c\x85&``\xb9\x86}\xeaf\x13\x1e\xf3F\x03\xac\x91\x07\x08\x13Cc\x90\x8fH\xfd>X6y\x1b[|\x14\xc4d\xaa[\xea\x80\xbb\x95S\x1f \xf5\x83I\xb7\x02B\xadm\xcb\xaf\x01\x19\x11\xd7\xd4\xee\xfe\xe7\xc1\x01\x05H\x8e4\x86y\xb1\xab\xd0\xec^\xfdm\xb43V\x10\xba\xb9\x85\xe0$3\x97/o\x8d\x11\xb4\x0c\x10\x04\xac,\x03\xb8\xf5r!w+\x95@\x88\xf3@6\xcc\x9d\xcf\xe2\x0c\xf6io+ \xca\xc1u][\xa9_@|\x85\x9a\x8d\xc51Y\x82xu/hW\xf0\x98\xa64\x18KG\xa8\xbe\xa4"l\xeb\x19\x96~\xedx\x94\xce\x0f\th[\xcc"\xf6\x8c\xcf\xdcxVB&"\xbb@\x106\x01\x08\xb90{\x06\x7f?S\xe8\x97\x1e\xd2\x11\xbc\x13\xd1\xe6\x1a\x08\x94\xfa\xb5RG\x1a\xc2\x06\xab\x04\x93e\xd7\x07\r\t \xf2y\x99\xf3\xe1q\x7f\xf7\x07\x9a\xf3>\x82\xc0\xb9\xa6\xbeu]r\x9f\xf02 \xd3&\x1ab\xed\xde\xc7@\xb4\xc7\'\x07\xb4\xc2x\x91\xb9\x08\x99l\xb4\xd1F\xe5#\xd2|\xf2\x06,\x03|\x82\x9c!\xef\x8aL#\xb3p+\xf5\x0bH\xfd\x9b\x8b\xbc\r\xbb>\x00\x12;\xee\xf9\xd6\x97\x8cI\x8e\xd0S\xafV\xae\x81ig\xef\x85\xd6\xf7\xdf\x7f\xff\xa2\r\xf6\x7f\xd5\x0e\x01\xd3\xc4\x110\x88\xff\xe4\'?in\xba\xe9\xa6\xe2y\x99\x94r\x028\x03&\xa4\x80\x04\xac2S~\xe4\xea\xdc\xbb\xeb\xfa\xd6\xd4\xf1\x18\x02\x10\xbb=\xfc\xce\xae\x8f,-\xc6\x18\xf0|\x8a\x8c\xed\x1e\x1c\x0c\xac.\xea\xa5~\xe1\xba\x9e\xb9\xcfkrd\xf3\xf3\x13K\x04\xe539+\x92\xdaJ\xf8\xb4A\xdb\x99m\xae\xb4\x9d\x92\xc6!\xe5\x98`\x92\x91_We\xce\x990s\x17Zc\x8e\x13\xed\x0b\x00(\xc0\'\xed^M\x1d\x01\x92\xd9\xaf^\xc0\x9b\xf0\xcb\xa3\xf6 \x89tJc\xe9\xc1\xc2\xea\xc6\xe55I\xc4\xd2\xb8]>\x9fR\x9b\xe5G\xe3\t\xd9d/\xfb\x87i\x85\xe3J+\xadT~\x1b\xc5x\xa2\xf7G\xd0\xc6\x18\x93K\xf3\x14\xf3\x1e\x00Z\xa7\xe1\xd9\xf1\xc8\xa2!\xad\x9a\x81]\x8f\xfcC\x1dkH<\n\x95\x03L\xd8\x9a\xb2k\x8ea\xbd\xa4\x95\xeb\xfb\x9f\'{Wk\xbdZ\xeb\x90<\x8e\xc2+\xb6\x12q\x93\x81\x82i\tS\xcdD\x9b\xb3\xc8\x0b @q}\x81\x10\x81\x122\r\xf3\xf3O&\x98Y\xb42\xee\x18\x87\x00\'o\xc6+\xcf\xe4Y\xc7V\xeah\x0cIx\x9b\x8d\x8c\n\xb3\xa9\xf1.r\xad\x955\xa6\xe6vy>oV\xb7\xd41\xf5\x90\x0e3E\xb4\x84\xa0\xf5\xe2\x80bW\xa2(\xb2^\xcft\x99\xc3\xe8\x9c<1\x80\xd4\xc2\x95f\xfe\xf6\xde{\xef2\xc0{\x07\x99q\x08\xf2\xe35\x80k\x07\x08\xae\xa9- H\x1a\xf2&A\xc2\x03&KB\r\x907\xe0\xd9B)>\xa4W8\xa7\xd2\xedX\x1e\xdc\xee\xde\xe7\xc9\xa9\x93w\xabs\xeaP\xd7G\x1e\x13F+~\xf9\x1c-&\x86k,\xc0\x98\xcd\xd44\x8b\xb9\xf6\xa5\x16\xe1\xc6+\xf3\x8c\xa3 #\x8d\xa0I\x004\xe1\xb45\x88\x13\xc4\x95\x96\x07\xd7@HG\xfe\xa1an\xb4#\x0fSc\xfey= \x0e\x8f\xf5\x90\xc1\xc0\xed\xea\xd6\xca\xf5\xe0oPg\xeb\t\xb8\x16\x14`\x8c\x17;\xee\xb8c1m,\x84\x1d\xee\xb4@G\r\x18X^\xa6)\xbf\xf2\xcd!0O\x11\r\xe0\x1a\xfbl\x9aVe\x1c\t\x108e`T\x00i\xbd\x88\xa4\xa3\xc2c#E\xe8\x11,rt\xee\xba\xfb:#\x8f\x8aF1C\xe6!v\xb8\x98\x8be\x16\xcf\x0c%?\x87@\x90\x91\x97%?\x93n\x06\xcf\x19\x00\xba\xfc\xf2\xd5\x14\xb9\xa7\x0e\x9f\x02$\x88\xd5T\xdf\x1f\x11\x0f&jW\xbfVn%\xd7\xd2{\x93&@\xbd\xdb\x9c\x83c#\xf8h\x1c\xe5U\xd2\x86\x1a\xd0\xb0e\x00[L\x8d9\xf2\x03FD\xd9\x120-\t\xe09zG\x9eE\x9f\xfa\xebU\x17S\xa1d@9\x1f\x11\x0f&jW\xbf\x9a\xb5\x910\xd2\xbbk B\xceq\x96\x89\x8d\r\x04l<11\xcc>0\xcfc\xcf*\x87sp\xef\xbd\xf7\x96\x9d\xf0\x9c\x86q\xff\xdf;5\xc0\x8b\x95q\x92j\x8d\n\xa7^\xe83c\x88\xf3\xd6Lc#\xa5}\xed\xdaY_\'d\x9eR\xbe\xfc56pk\xfd\x9b\x83\x9f\x84\x8a\xc7%\x1f\xd6\xe3\xad\x99\xf8\xd5R\xb1-\x80\x00\xc6\xc7\xa9f\xf5@$\xdf\xba\x13\x84Q\xf1\xb2\xea\x0b\xa85S}>6p\x04"\x1d\xaa\xd3H\x9eh\x8f\xb1\x81\xc9\xc9\x16\xa0Dw\r\xe0V\r=\x1b\x01c\x03<Sg\xc6o\xdc\x91\xdf\xe4\xdaO\xecr$\x00\x18\x10[\xe5_4$\x05\xe5b2\xe4\xbc?\xaa\xf3\x0f\x06\xee\x8f\xb4\xb5\x15\x90Z0\xe1\xc8\x050\x06e\x91\\\xae,\x01\x1b\xb0-P1g\xccZm\xba<\xe3\x9aO\'h\x86P=-\xb1o\xccO\x82\xf0\xd2\xda\xbd\x0f\x951$\x85\xe4bk\xa6\xb1\x8d\xea\xb69\xd6\xc2A9\x07Z\xec=\xb7\x98\xbb,\x96\xc5\x9d\x05\x8a\xa0+7\xd8u\xa6*\x1a\x85\x02\xa2\xc9\xa45}\x83\xbb\x19?\x87\xc0\xc6\xf0\x94\x9f\x8e\x91w\xf7\r\xea\xb9\x10\xaa3\x8d\xcd\xa4\x8dugD9w\x8c\x90\xa5\xd9\x7f\xab\x82&\x8bf\xff\x84l\xe5\xd0/\xb5\xe6\xd7\x80"`G \x1agL\xac\x8d;\x9e\x11\xa04\xbe\xb8\xa7l\xa6K\xfe\xbc\x7f\xe8\x0f\xee\x87C\x01$B\x96\xa6\x05\xe2\\~#\x8b\xe7\x04\x10\xe1y+\x8f\xb6\x99\x1a;\xe4\xcbs\x04\x0e(\x1b\xc3-\x1d\xd3*\x1e\x9a\x95I\x03\x7f\xf2\r\x01\xd2!E`\xb5\x90yU\xc6\x86\xf4z\xa0\xf8\xce\xd0.F\xe1\x16\xf92`G\x03\xfcV>\x00yh\x02\x96\xc2\xf5\xc9\x8b\x87\x00\xe9\x90\xa2%\x01#i\xa6\xcbN\x14\x9e\x96^oL\xf1IB\xb6\x02\x05\x08\xcf\x98(\xfaIZ\x9a\x04\x90|Kb\xc7\xa5\xfbx\x08\x90\x0e(B\xc2\xb5\xe0\x08\x1a(6<l\xbe\xf9\xe6%\x1aL\xd0\xb6\x0cqs-|\x99\x1c2_"\xc3\xdce\xa1\x17\xee2m\xf2#5\xb6\xbe\x02N9\xc0\xcb{\xd0\x10 \xc3!\xc2\x8a\xa0\xeat\xc6\t\x81W\x91pn0A\x1b\xb0\xad\xbb\xfb1f\xdb\x87\xb8\xc3\xf6j\x89\x90\x8b\xfc\x8a\x00\x1bG\xac\xb9[q\r\x18u\xd9h\x08\x90\x11P\xad\x19\x04gPwD\xae\xd9\x07\xec?\xb9|/\x0f\x14\x82\x17\xf3\xb2\x16o\\\x11\xf55\x07\x01\x86y\x0b\xed\x10\xac\x14\xaaW.m\x1b\xd2\x90.\x88\xd0\xd2\x8b1\xed\x88\xe0\xdcc\x9a\xec0\xd9y\xe7\x9d\xcb\\C\x98\x040\xccS\x16\xc5\x983\xe7\xc2(\xa2\xbe\xc20\x9e\xa3a\xb5v\x0c\x01\xd2\x0f\x11\x10\x81e\x80\x0e(\xb5\x10\xa5\r\xf0\xe6\x1a\xcc\x93\x1f\xe2d\xb6,\x1d\xf3\xc0\xac\xa3\x18\xc4\xad\xdf\xdb\xcff\x0f\x18 Rf\x06\xfe!@:\xa0\x08)\x82\x8f\xc0r-D\xa0\xc2$\xc2\xf1\xf6\x13\x9bc\x98w\xf8\r\x16\xb3t\x93FZ\xc4\xab\x8a\xc9K\x19\x9e\xc55\r\x012\x00D\xc8z;`\x08^\x98\xdd\x04\xd2\\\xc3x\x91u\x10\xf9\xc2\x01\xa5\x06\x17\r\x012\x8aD\xb8\x11v\x04\\\x0b\xdd\xbd\x08\xbd5o;\x1a\x02d\x14\xa9U\xc8\xad\x80\x0c\xef:\xca\xf5\x9a\x86\x00\x19\x05\x8a@k\x01\x13\xb8A;\x9a1<\x87\xa0\xce_\xd3\x10 \xa3@\x04\x1f\xb78\xc2\x05D4\x06\xd5\xc2O\x9e\xe4oGC\x80|N\x14\xe1\xb7R\xae\xb7\x07\xa4i\xfe\x0f\xcd\xbf\xef\x8f\x8e\xf0\xe4\xeb\x00\x00\x00\x00IEND\xaeB`\x82'
        # import tempfile
        # with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as f:
        #     f.write(binary_data)
        #     f.flush()  # 确保写入磁盘
        #     icon = QIcon(f.name)
        # self.setWindowIcon(icon)

        window_icon = QIcon(f"{os.path.dirname(sys.argv[0])}/_internal/aaa_etc/weixinshoucang.ico")
        # 设置窗口图标
        self.setWindowIcon(window_icon)
        # self.setWindowIcon(FluentIcon.GAME.icon())

        # 初始化数据库配置连接信息
        data_dict = self.cf_db.read()
        data_dict_key = sorted(data_dict.keys(), key=str.lower)
        self.myUI.lswDB.addItems(data_dict_key)
        # 初始化数据库连接选择
        self.myUI.cbb_choose_db.clear()
        self.myUI.cbb_choose_db.addItems(sorted(data_dict.keys(), key=str.lower))
        self.myUI.cbb_choose_db.setCurrentIndex(-1)
        # 表、过程清单设置列宽
        self.myUI.tbw_table.setColumnWidth(0, 120)
        self.myUI.tbw_table.setColumnWidth(1, 200)
        self.myUI.tbw_table.setColumnWidth(2, 200)
        self.myUI.tbw_procedure.setColumnWidth(0, 120)
        self.myUI.tbw_procedure.setColumnWidth(1, 150)
        self.myUI.tbw_procedure.setColumnWidth(2, 200)

        self.myUI.splitter_table.setSizes([1000, 2000, 0])  # 设置每个的宽度，单位像素，受外部框体的限制
        self.myUI.splitter_procedure.setSizes([1000, 2000, 0])  # 设置每个的宽度，单位像素，受外部框体的限制

        self.myUI.splitter_procedure.setSizes([2000, 4000, 1000])  # 设置每个的宽度，单位像素，受外部框体的限制

        # 读取sql标签列表
        file_list = [file.split(".")[0] for file in os.listdir(self.sql_book_dir) if file.endswith(".txt")]
        file_list.sort()  # 升序
        self.myUI.lsw_book.addItems(file_list)
        # 读取dpi设置
        self.cf_dpi = ConfigFile(os.path.dirname(sys.argv[0]) + "/_internal/aaa_etc/environ.ini", "utf=8")  # 配置文件路径：当前
        self.myUI.cbb_dpi.setCurrentIndex(-1)
        dpi_value = self.cf_dpi.read()['environment']['qt_font_dpi']
        self.myUI.cbb_dpi.setText(dpi_value + ("" if dpi_value == "AUTO" else "%"))
        # 设置table、procedure表头
        self.myUI.tbw_table.horizontalHeader().setVisible(True)
        self.myUI.tbw_procedure.horizontalHeader().setVisible(True)

        # 读取结构转化清单
        self.menu1 = RoundMenu(parent=self)
        self.menu1.addAction(Action('to_oracle'))
        self.menu1.addAction(Action('to_gbase'))
        self.menu1.addAction(Action('to_dameng'))
        self.myUI.btn_trans.setMenu(self.menu1)

        # 设置图片
        self.myUI.btn_connect.setIcon(FluentIcon.CONNECT)
        self.myUI.btn_get_data.setIcon(FluentIcon.SEARCH)
        self.myUI.btn_reset.setIcon(FluentIcon.CANCEL)
        self.myUI.btn_save_db.setIcon(FluentIcon.SAVE)
        self.myUI.btn_test_db.setIcon(FluentIcon.CONNECT)
        self.myUI.btn_sql_execute_1.setIcon(FluentIcon.PLAY)
        self.myUI.btn_sql_execute_2.setIcon(FluentIcon.PLAY)
        self.myUI.btn_sql_execute_3.setIcon(FluentIcon.PLAY)
        self.myUI.btn_sql_save_1.setIcon(FluentIcon.QUICK_NOTE)
        self.myUI.btn_sql_save_2.setIcon(FluentIcon.HEART)
        self.myUI.btn_sql_save_3.setIcon(FluentIcon.HEART)
        self.myUI.btn_sql_del_1.setIcon(FluentIcon.DELETE)
        self.myUI.btn_sql_del_2.setIcon(FluentIcon.DELETE)
        self.myUI.btn_sql_del_3.setIcon(FluentIcon.DELETE)
        self.myUI.tbw_column.btn_export.setIcon(FluentIcon.DOWNLOAD)
        self.myUI.tbw_sql_result_1.btn_export.setIcon(FluentIcon.DOWNLOAD)
        self.myUI.tbw_sql_result_2.btn_export.setIcon(FluentIcon.DOWNLOAD)
        self.myUI.tbw_sql_result_3.btn_export.setIcon(FluentIcon.DOWNLOAD)
        self.myUI.btn_dbinfo_add.setIcon(FluentIcon.ADD)
        self.myUI.btn_dbinfo_del.setIcon(FluentIcon.REMOVE)

        # 图标
        self.myUI.icon_dpi.setIcon(FluentIcon.ZOOM)
        self.myUI.icon_yunmu.setIcon(FluentIcon.TRANSPARENT)
        self.myUI.icon_theme.setIcon(FluentIcon.BRUSH)
        self.myUI.icon_theme_color.setIcon(FluentIcon.PALETTE)
        self.myUI.btn_choose_color.clicked.connect(
            lambda: ColorDialog(QColor(0, 159, 170), 'choose locor', self, True).exec())
        self.myUI.icon_ykl.setIcon(FluentIcon.ALBUM)
        self.myUI.icon_help.setIcon(FluentIcon.HELP)
        self.myUI.icon_res.setIcon(FluentIcon.FEEDBACK)
        self.myUI.icon_update.setIcon(FluentIcon.DOCUMENT)
        self.myUI.icon_about.setIcon(FluentIcon.INFO)
        self.myUI.icon_file_blood.setIcon(FluentIcon.APPLICATION)
        self.myUI.icon_col_trans.setIcon(FluentIcon.APPLICATION)
        self.myUI.icon_similarity.setIcon(FluentIcon.APPLICATION)

        # 导出MAPPING的MENU
        self.menu_export_database = RoundMenu(parent=self)  # 菜单
        self.action_export_etl = Action('从etl库', self)  # 动作etl
        self.action_export_etl.triggered.connect(self.export_mapping)  # 动作绑定函数
        self.action_export_etl_bl = Action('从etl_bl库', self)  # 动作etl_bl
        self.action_export_etl_bl.triggered.connect(self.export_mapping)  # 动作绑定函数
        self.menu_export_database.addAction(self.action_export_etl)  # 动作插入到菜单
        self.menu_export_database.addAction(self.action_export_etl_bl)  # 动作插入到菜单
        self.myUI.btn_mapping.setMenu(self.menu_export_database)  # 按钮设置菜单

        # 读取MAPPING的MENU
        self.menu_load_database = RoundMenu(parent=self)  # 菜单
        self.action_load_etl = Action('从etl库', self)  # 动作etl
        self.action_load_etl.triggered.connect(self.get_datamapping)  # 动作绑定函数
        self.action_load_etl_bl = Action('从etl_bl库', self)  # 动作etl_bl
        self.action_load_etl_bl.triggered.connect(self.get_datamapping)  # 动作绑定函数
        self.menu_load_database.addAction(self.action_load_etl)  # 动作插入到菜单
        self.menu_load_database.addAction(self.action_load_etl_bl)  # 动作插入到菜单
        self.myUI.btn_load_mapping.setMenu(self.menu_load_database)  # 按钮设置菜单

        # 写入MAPPING的MENU
        self.menu_upload_database = RoundMenu(parent=self)
        self.action_upload_etl = Action('到etl库', self)
        self.action_upload_etl.triggered.connect(self.update_datamapping)
        self.action_upload_etl_bl = Action('到etl_bl库', self)
        self.action_upload_etl_bl.triggered.connect(self.update_datamapping)
        self.menu_upload_database.addAction(self.action_upload_etl)
        self.menu_upload_database.addAction(self.action_upload_etl_bl)
        # self.copy_action.triggered.connect(self.copy_selection)
        self.myUI.btn_upload_mapping.setMenu(self.menu_upload_database)

        # self.dropDownPushButton1.setMenu(self.menu)

        # 窗口左上角
        self.setGeometry(0, 0, 1000, 600)

        # 版本
        self.myUI.label_24.setText("版权@2024,LJZ  版本2025.4.23")

        # 日志文件
        logger.add(os.path.dirname(sys.argv[0]) + "/_internal/aaa_log/loguru.log")

    def load_tablewidget(self, widget: QTableWidget, head: list = None, data: list = None):
        """ 向widget加载数据 """
        if head:
            widget.setColumnCount(len(head))
            widget.setHorizontalHeaderLabels(head)
        if data:
            widget.setRowCount(len(data))
            # 加工表清单
            for row in range(len(data)):
                for col in range(len(data[0])):
                    item = QTableWidgetItem(str(data[row][col] or ""))
                    widget.setItem(row, col, item)
            # i = 0
            # for row in range(len(data)):
            #     for col in range(len(data[0])):
            #         item = QTableWidgetItem(str(data[row][col] or ""))
            #         widget.setItem(row, col, item)
            #     i += 1
            #     if i % 500 == 500:
            #         sleep(1)

        else:
            widget.setRowCount(0)

    def resize_columns_to_content(self):
        pass

    def save_db_info(self):
        """保存数据库连接"""
        # 如果没选中任何连接，则不执行保存
        if not self.myUI.lswDB.selectedItems():
            self.createInfoInfoBar("未选中任何连接")
            return

        # 获取连接名
        old_db_name = self.myUI.lswDB.currentItem().text()
        # 获取文本框内容
        name = self.myUI.edt_name.text()
        desc = self.myUI.edt_desc.text()
        db_type = self.myUI.cbb_type.currentText()
        host = self.myUI.edt_host.text()
        port = self.myUI.edt_port.text()
        database = self.myUI.edt_database.text()
        username = self.myUI.edt_username.text()
        password = self.myUI.edt_password.text()
        charset = self.myUI.edt_charset.text()
        data_dict = {"name": name,
                     "desc": desc,
                     "type": db_type,
                     "host": host,
                     "port": port,
                     "database": database,
                     "username": username,
                     "password": password,
                     "charset": charset}
        # 执行保存
        self.cf_db.update(old_db_name, data_dict, name)
        # 显示结果
        self.createSuccessInfoBar("已保存")
        # 更新数据库配置项连接项名
        self.myUI.lswDB.currentItem().setText(name)
        # 更新选择数据库连接项名
        for i in range(self.myUI.cbb_choose_db.count()):
            if self.myUI.cbb_choose_db.itemText(i) == old_db_name:
                self.myUI.cbb_choose_db.setItemText(i, name)
                break

    def save_sql_book(self):
        text = self.sender().objectName()
        if text == "btn_sql_save_1":
            book_name = self.myUI.edt_sql_book_1.text()
            if not book_name:
                self.createInfoInfoBar("未输入书签名")
                return
            sql_body = self.myUI.edt_sql_body_1.toPlainText()

        if text == "btn_sql_save_2":
            book_name = self.myUI.edt_sql_book_2.text()
            if not book_name:
                self.createInfoInfoBar("没有输入书签名")
                return
            sql_body = self.myUI.edt_sql_body_2.toPlainText()

        if text == "btn_sql_save_3":
            book_name = self.myUI.edt_sql_book_3.text()
            if not book_name:
                self.createInfoInfoBar("没有输入书签名")
                return
            sql_body = self.myUI.edt_sql_body_3.toPlainText()

        with open(f"_internal/aaa_book/{book_name}.txt", "w") as f:
            f.write(sql_body)
        book_list = [self.myUI.lsw_book.item(row).text() for row in range(self.myUI.lsw_book.count())]
        if book_name not in book_list:
            self.myUI.lsw_book.addItem(book_name)
            book_list.append(book_name)
        self.myUI.lsw_book.setCurrentRow(book_list.index(book_name))  # 显示当前标签
        self.createSuccessInfoBar("已保存")

    def set_tbw_current(self):
        """设置当前选中的tbw"""
        self.tbw_current = self.sender()
        # 解决2个tablewidget快捷键冲突
        if self.tbw_current == self.myUI.tbw_table:
            self.myUI.tbw_table.addAction(self.myUI.tbw_table.copy_action)
            self.myUI.tbw_column.tbw_table.removeAction(self.myUI.tbw_column.tbw_table.copy_action)
        if self.tbw_current == self.myUI.tbw_column.tbw_table:
            self.myUI.tbw_column.tbw_table.addAction(self.myUI.tbw_column.tbw_table.copy_action)
            self.myUI.tbw_table.removeAction(self.myUI.tbw_table.copy_action)

    def show_db_info(self):
        """将选中的连接显示到页面"""
        db_name = self.myUI.lswDB.currentItem().text()
        # 从配置文件读取选中的连接
        data_dict = self.cf_db.read()
        db_info = data_dict[db_name]
        type_index = {'': 0,
                      'oracle': 1,
                      'mysql': 2,
                      'gbase': 3,
                      'dameng': 4
                      }[db_info["type"].lower()]

        # 显示到编辑框
        self.myUI.edt_name.setText(db_name)
        self.myUI.edt_desc.setText(db_info["desc"])
        self.myUI.cbb_type.setCurrentIndex(type_index)
        self.myUI.edt_host.setText(db_info["host"])
        self.myUI.edt_port.setText(db_info["port"])
        self.myUI.edt_database.setText(db_info["database"])
        self.myUI.edt_username.setText(db_info["username"])
        self.myUI.edt_password.setText(db_info["password"])
        self.myUI.edt_charset.setText(db_info["charset"])

    def show_procedure_blood(self):
        """展示过程体血缘"""
        text = self.myUI.edt_procedure_body.toPlainText()
        tab_list = pars_text(text)
        bloods = "\n".join(tab_list)
        self.myUI.edt_procedure_blood.setText(bloods)

    def show_sql_body(self):
        """显示书签内容"""
        item = self.myUI.lsw_book.currentItem()
        file_name = self.sql_book_dir + item.text() + ".txt"
        # 获取sql_body
        with open(file_name, 'rb') as f:  # 获取文件编码
            data = f.read()
            file_encoding = chardet.detect(data)["encoding"]
        with open(file_name, encoding=file_encoding) as f:
            sql_body = f.read()
        if self.myUI.segw_sql.currentRouteKey() == "tab_sql_1":  # 显示到sql编辑框1
            self.myUI.edt_sql_book_1.setText(item.text())
            self.myUI.edt_sql_body_1.setText(sql_body)
        if self.myUI.segw_sql.currentRouteKey() == "tab_sql_2":  # 显示到sql编辑框2
            self.myUI.edt_sql_book_2.setText(item.text())
            self.myUI.edt_sql_body_2.setText(sql_body)
        if self.myUI.segw_sql.currentRouteKey() == "tab_sql_3":  # 显示到sql编辑框3
            self.myUI.edt_sql_book_3.setText(item.text())
            self.myUI.edt_sql_body_3.setText(sql_body)

    def sort_result(self, index):
        """点击表头，排序功能"""
        tbw: QTableWidget = self.sender().parent()  # 获取点击表头的父类
        # 获取结果集数据
        row_data = []
        for row in range(tbw.rowCount()):
            col_data = []
            for col in range(tbw.columnCount()):
                col_data.append(tbw.item(row, col).text())
            row_data.append(col_data)
        # 获取当前排序标志
        sort_flag = True
        if tbw == self.myUI.tbw_column:
            self.tbw_column_sort_flag = not self.tbw_column_sort_flag
            sort_flag = self.tbw_column_sort_flag
        if tbw == self.myUI.tbw_sql_result_1.tbw_table:
            self.tbw_sql_result_1_sort_flag = not self.tbw_sql_result_1_sort_flag
            sort_flag = self.tbw_sql_result_1_sort_flag
        if tbw == self.myUI.tbw_sql_result_2.tbw_table:
            self.tbw_sql_result_2_sort_flag = not self.tbw_sql_result_2_sort_flag
            sort_flag = self.tbw_sql_result_2_sort_flag
        if tbw == self.myUI.tbw_sql_result_3.tbw_table:
            self.tbw_sql_result_3_sort_flag = not self.tbw_sql_result_3_sort_flag
            sort_flag = self.tbw_sql_result_3_sort_flag
        # 判断选中列是否全是数字
        decimal_flag = True  # 默认全数字
        for row in row_data:
            if not re.match("^[+-]?[0-9]+[.]?[0-9]*$", row[index]):
                decimal_flag = False  # 不是全数字
        if decimal_flag:
            row_data.sort(key=lambda x: float(x[index]), reverse=sort_flag)  # decimal排序 , reverse=True降序
            self.createInfoInfoBar("按数字降序排序" if sort_flag else "按数字升序排序")
        else:
            row_data.sort(key=lambda x: x[index], reverse=sort_flag)  # 字符排序
            self.createInfoInfoBar("按字符降序排序" if sort_flag else "按字符升序排序")
        # 显示到表单
        for row in range(tbw.rowCount()):
            for col in range(tbw.columnCount()):
                tbw.item(row, col).setText(row_data[row][col])

    def test_db_connect(self):

        # 前端校验
        creator = self.myUI.cbb_type.currentText()
        host = self.myUI.edt_host.text()
        port = self.myUI.edt_port.text()
        user = self.myUI.edt_username.text()
        password = self.myUI.edt_password.text()
        schema = self.myUI.edt_database.text()
        if not (creator and host and port and user and password and schema):
            self.createInfoInfoBar("信息项不完整")
            return

        # ui
        self.myUI.btn_test_db.setEnabled(False)

        cfg = {'type': self.myUI.cbb_type.currentText().lower(),
               'host': self.myUI.edt_host.text(),
               'port': int(self.myUI.edt_port.text()),
               'user': self.myUI.edt_username.text(),
               'password': self.myUI.edt_password.text(),
               'schema': self.myUI.edt_database.text()}
        self.thread_test_pool = PoolThread(**cfg)

        # 信号绑定函数
        self.thread_test_pool.sign_end.connect(lambda: self.createSuccessInfoBar("连接成功"))
        self.thread_test_pool.sign_end.connect(lambda: self.myUI.btn_test_db.setEnabled(True))
        self.thread_test_pool.sign_err.connect(self.createErrorInfoBar)
        self.thread_test_pool.sign_err.connect(logger.warning)
        self.thread_test_pool.sign_err.connect(lambda: self.myUI.btn_test_db.setEnabled(True))

        # 10秒连接超时
        QTimer.singleShot(10000, lambda: self.timer_out(self.thread_test_pool, "连接超时"))
        self.thread_test_pool.start()

    def test_func(self, *args, **kwargs):
        print(args)
        print(kwargs)

    def timer_out(self, thread: PoolThread, text):
        if thread.isRunning():
            thread.sign_err.emit(text)
            thread.terminate()
            thread.wait()
        # 定时器正常停止
        # self.createInfoInfoBar("定时器正常停止")

    def update_datamapping(self):
        """ 同步datamapping到数据库 """
        if not self.myUI.tbw_table.currentItem():
            self.createInfoInfoBar("未选中表")
            return
        if self.myUI.tbw_column.tbw_table.rowCount() == 0:  # 如果没有数据则不操作
            self.createInfoInfoBar("无数据")
            return
        database_name = re.findall(r"[a-zA-Z_]+", self.sender().text())[0]
        item = self.myUI.tbw_table.currentItem()
        table_name = self.myUI.tbw_table.item(item.row(), 1).text()

        # 拼接sql
        # delete from etl.datamapping where (seq_num='1' and t_tab_eng_name='tab') or (seq_num='1' and t_tab_eng_name='tab')
        sql_delete = f"delete from {database_name}.datamapping where "
        values = ""
        # for row in range(5): # 测试用
        for row in range(self.myUI.tbw_column.tbw_table.rowCount()):
            value = '('
            for col in range(self.myUI.tbw_column.tbw_table.columnCount()):
                item = self.myUI.tbw_column.tbw_table.item(row, col)
                text = item.text().replace("'", "''") if item else ''
                value += f"'{text}',"
            value = value[:-1] + '),'
            values += value

            sql_delete += f"(seq_num='{self.myUI.tbw_column.tbw_table.item(row, 0).text()}' and t_tab_eng_name='{self.myUI.tbw_column.tbw_table.item(row, 1).text()}') or "  # delete sql

        values = values[:-1]  # 去掉结尾的,
        sql_delete = sql_delete[:-4]  # 去掉结尾的 or
        sql = f"insert into {database_name}.datamapping values{values}"

        connection = self.thread_pool.get_connection()
        cursor = connection.cursor()
        try:

            # cursor.execute(sql) # 1.尝试插入，确保能插入
            # cursor.execute(f"delete from {database_name}.datamapping where lower(t_tab_eng_name)=lower('{table_name}')") # 2.delete
            logger.info(sql_delete)
            cursor.execute(sql_delete)  # 2.delete
            logger.info(sql)
            cursor.execute(sql)  # 3.insert
            connection.commit()  # 提交
            self.createSuccessInfoBar(f"已同步到{database_name}.datamapping")
        except Exception as e:
            logger.error(f"sql异常{e}")
            self.createErrorInfoBar(f"sql异常{e}")
        finally:
            connection.close()


if __name__ == "__main__":

    # pyinstaller -D -i C:\Users\lojn\PycharmProjects\DataView\img\weixinshoucang.ico --add-data drivers/dameng/dpi/*:. --add-data drivers/oracle/instantclient/*.dll:. --add-data _internal/aaa_book/*:aaa_book --add-data _internal/aaa_etc/*:aaa_etc --add-data _internal/aaa_sql/*:aaa_sql -n DataView App.py -w


    # 环境变量 dpi
    cf = ConfigFile(os.path.dirname(sys.argv[0]) + "/_internal/aaa_etc/environ.ini", "utf=8")  # 配置文件路径：当前
    qt_font_dpi = cf.read()["environment"]["qt_font_dpi"]
    if qt_font_dpi != "AUTO":
        os.environ["QT_FONT_DPI"] = qt_font_dpi
    app = QtWidgets.QApplication(sys.argv)
    ex = MyMainWindow()  # 示例窗口
    ex.show()  # 显示
    ex.check_enable()  # 检查可用性
    # 建议使用os.path.realpath(os.path.dirname(sys.argv[0]))
    # sys.exit(app.exec_()) pyqt5
    sys.exit(app.exec())  # pyside6

    # self.menu = RoundMenu(parent=self)
    # self.menu.addAction(Action(FIF.BASKETBALL, 'Basketball'))
    # self.menu.addAction(Action(FIF.ALBUM, 'Sing'))
    # self.menu.addAction(Action(FIF.MUSIC, 'Music'))
    # self.dropDownPushButton1.setMenu(self.menu)
